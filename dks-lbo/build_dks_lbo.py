"""
Dick's Sporting Goods LBO Model Builder
Generates a portfolio-quality LBO model as DKS_LBO_Model.xlsx
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.views import SheetView
import os

OUTPUT_PATH = "/Users/paco/Downloads/DKS_LBO_Model.xlsx"

# ── Color constants ──────────────────────────────────────────────────────────
BLUE_FILL      = "DBE5F1"   # input cells
DARK_BLUE_FONT = "17375E"
GREEN_FILL     = "EBF1DE"   # cross-sheet links
DARK_GREEN_FONT= "375623"
NAVY_FILL      = "1F3864"   # main section headers
NAVY2_FILL     = "2F5496"   # sub-headers / year headers
GRAY_FILL      = "F2F2F2"   # total / subtotal rows
WHITE          = "FFFFFF"
BLACK          = "000000"

# ── Style helpers ────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=BLACK, bold=False, italic=False, size=10, name="Calibri"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _bottom_border():
    s = Side(style="thin")
    return Border(bottom=s)

def _thick_border():
    thin = Side(style="thin")
    thick = Side(style="medium")
    return Border(left=thick, right=thick, top=thick, bottom=thick)

# Number formats
FMT_COMMA0  = '#,##0'
FMT_COMMA1  = '#,##0.0'
FMT_PCT1    = '0.0%'
FMT_PCT2    = '0.00%'
FMT_MULT    = '0.0"x"'
FMT_GENERAL = 'General'

def style_input(ws, row, col, value=None, fmt=FMT_COMMA0, h_align="right"):
    """Blue fill, dark blue font — hardcoded input cell."""
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.fill = _fill(BLUE_FILL)
    c.font = _font(DARK_BLUE_FONT)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def style_formula(ws, row, col, formula=None, fmt=FMT_COMMA0,
                  h_align="right", bold=False, fill_hex=None):
    """Black font formula cell."""
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.font = _font(BLACK, bold=bold)
    c.alignment = _align(h_align)
    c.number_format = fmt
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c

def style_link(ws, row, col, formula=None, fmt=FMT_COMMA0, h_align="right"):
    """Light green fill, dark green font — cross-sheet link."""
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.fill = _fill(GREEN_FILL)
    c.font = _font(DARK_GREEN_FONT)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def style_header_main(ws, row, col, label, col_span=1):
    """Dark navy fill, white bold font — main section header."""
    c = ws.cell(row=row, column=col, value=label)
    c.fill = _fill(NAVY_FILL)
    c.font = _font(WHITE, bold=True)
    c.alignment = _align("left")
    if col_span > 1:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=row, end_column=col + col_span - 1
        )
    return c

def style_header_sub(ws, row, col, label, h_align="center"):
    """Medium blue fill, white font — sub-header / year header."""
    c = ws.cell(row=row, column=col, value=label)
    c.fill = _fill(NAVY2_FILL)
    c.font = _font(WHITE, bold=True)
    c.alignment = _align(h_align)
    return c

def style_label(ws, row, col, label, bold=False, indent=0):
    """Plain label cell (no fill, black font, left-aligned)."""
    c = ws.cell(row=row, column=col, value=label)
    c.font = _font(BLACK, bold=bold)
    c.alignment = Alignment(horizontal="left", vertical="center",
                            indent=indent)
    return c

def style_total(ws, row, col, formula=None, fmt=FMT_COMMA0,
                h_align="right", bold=True):
    """Gray fill total row."""
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.fill = _fill(GRAY_FILL)
    c.font = _font(BLACK, bold=bold)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def no_gridlines(ws):
    ws.sheet_view.showGridLines = False

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def freeze(ws, cell_ref):
    ws.freeze_panes = cell_ref


# ════════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_assumptions(ws):
    no_gridlines(ws)
    ws.sheet_properties.tabColor = "2E75B6"

    set_col_width(ws, 1, 36)
    set_col_width(ws, 2, 16)
    for c in range(3, 8):
        set_col_width(ws, c, 12)

    # ── TRANSACTION ASSUMPTIONS ──────────────────────────────────────────────
    style_header_main(ws, 3, 1, "TRANSACTION ASSUMPTIONS", col_span=7)

    rows_ta = [
        (4,  "LTM Revenue ($M)",                    13000,   None,         FMT_COMMA0),
        (5,  "LTM EBITDA ($M)",                     1700,    None,         FMT_COMMA0),
        (6,  "LTM EBITDA Margin",                   None,    "=B5/B4",     FMT_PCT1),
        (7,  "Entry EV / LTM EBITDA Multiple",      9.0,     None,         FMT_MULT),
        (8,  "Transaction Enterprise Value ($M)",   None,    "=B5*B7",     FMT_COMMA0),
        (9,  "Existing Net Debt at Close ($M)",     300,     None,         FMT_COMMA0),
        (10, "Transaction Equity Value ($M)",        None,    "=B8-B9",     FMT_COMMA0),
        (11, "Transaction Fees (% of TEV)",         0.02,    None,         FMT_PCT1),
        (12, "Transaction Fees ($M)",               None,    "=B8*B11",    FMT_COMMA0),
    ]
    for (r, lbl, inp, frm, fmt) in rows_ta:
        style_label(ws, r, 1, lbl, indent=1)
        if inp is not None:
            style_input(ws, r, 2, inp, fmt=fmt)
        else:
            style_formula(ws, r, 2, frm, fmt=fmt)

    # Row 13: Total Uses — bold total row
    style_label(ws, 13, 1, "Total Uses ($M)", bold=True, indent=1)
    style_total(ws, 13, 2, "=B10+B12", fmt=FMT_COMMA0)

    # ── SOURCES OF FUNDS ─────────────────────────────────────────────────────
    style_header_main(ws, 15, 1, "SOURCES OF FUNDS", col_span=7)

    rows_sf = [
        (16, "Term Loan B — Leverage Multiple (x LTM EBITDA)", 5.5,  None,       FMT_MULT),
        (17, "Term Loan B ($M)",                               None,  "=B5*B16",  FMT_COMMA0),
        (18, "Revolving Credit Facility — Capacity ($M)",       750,  None,       FMT_COMMA0),
        (19, "Revolving Credit Facility — Drawn at Close ($M)", 0,    None,       FMT_COMMA0),
        (20, "Total Debt at Close ($M)",                        None,  "=B17+B19", FMT_COMMA0),
        (21, "Sponsor Equity ($M)",                             None,  "=B13-B20", FMT_COMMA0),
    ]
    for (r, lbl, inp, frm, fmt) in rows_sf:
        style_label(ws, r, 1, lbl, indent=1)
        if inp is not None:
            style_input(ws, r, 2, inp, fmt=fmt)
        else:
            style_formula(ws, r, 2, frm, fmt=fmt)

    style_label(ws, 22, 1, "Total Sources ($M)", bold=True, indent=1)
    style_total(ws, 22, 2, "=B20+B21", fmt=FMT_COMMA0)

    style_label(ws, 23, 1, "Leverage at Close (x LTM EBITDA)", indent=1)
    style_formula(ws, 23, 2, "=B20/B5", fmt=FMT_MULT)

    style_label(ws, 24, 1, "Equity Contribution (%)", indent=1)
    style_formula(ws, 24, 2, "=B21/B13", fmt=FMT_PCT1)

    # ── DEBT TERMS ───────────────────────────────────────────────────────────
    style_header_main(ws, 26, 1, "DEBT TERMS", col_span=7)

    rows_dt = [
        (27, "Base Rate (SOFR)",                      0.045,   None,           FMT_PCT2),
        (28, "Term Loan B — Credit Spread",           0.035,   None,           FMT_PCT2),
        (29, "Term Loan B — All-In Interest Rate",    None,    "=B27+B28",     FMT_PCT2),
        (30, "Term Loan B — Annual Amortization (%)", 0.01,    None,           FMT_PCT1),
        (31, "Revolving Credit Facility — Credit Spread", 0.0275, None,        FMT_PCT2),
        (32, "Revolving Credit Facility — All-In Rate", None,  "=B27+B31",     FMT_PCT2),
        (33, "Undrawn Commitment Fee (%)",            0.00375, None,           FMT_PCT2),
        (34, "Cash Tax Rate",                         0.25,    None,           FMT_PCT1),
    ]
    for (r, lbl, inp, frm, fmt) in rows_dt:
        style_label(ws, r, 1, lbl, indent=1)
        if inp is not None:
            style_input(ws, r, 2, inp, fmt=fmt)
        else:
            style_formula(ws, r, 2, frm, fmt=fmt)

    # ── OPERATING ASSUMPTIONS ────────────────────────────────────────────────
    # Row 36: section label in col A (no merge), year sub-headers in cols C-G
    c36 = ws.cell(row=36, column=1, value="OPERATING ASSUMPTIONS")
    c36.fill = _fill(NAVY_FILL)
    c36.font = _font(WHITE, bold=True)
    c36.alignment = _align("left")
    # Fill col B with navy too
    c36b = ws.cell(row=36, column=2)
    c36b.fill = _fill(NAVY_FILL)
    # Year sub-headers in cols C-G (not merged cells)
    for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
        style_header_sub(ws, 36, 3+i, yr)

    # CHANGE 1: Updated Revenue Growth and EBITDA Margin assumptions
    oa_inputs = [
        (37, "Revenue Growth",                  [0.04, 0.04, 0.05, 0.05, 0.05], FMT_PCT1),
        (38, "EBITDA Margin",                   [0.140, 0.150, 0.155, 0.160, 0.165], FMT_PCT1),
        (39, "D&A (% of Revenue)",              [0.027, 0.027, 0.027, 0.027, 0.027], FMT_PCT1),
        (40, "Capital Expenditures (% of Revenue)", [0.035, 0.035, 0.035, 0.035, 0.035], FMT_PCT1),
        (41, "Change in NWC (% of Δ Revenue)",  [0.05, 0.05, 0.05, 0.05, 0.05], FMT_PCT1),
    ]
    for (r, lbl, vals, fmt) in oa_inputs:
        style_label(ws, r, 1, lbl, indent=1)
        for i, v in enumerate(vals):
            style_input(ws, r, 3+i, v, fmt=fmt)

    # ── EXIT ASSUMPTIONS ─────────────────────────────────────────────────────
    style_header_main(ws, 43, 1, "EXIT ASSUMPTIONS", col_span=7)

    style_label(ws, 44, 1, "Exit Year", indent=1)
    style_input(ws, 44, 2, 5, fmt=FMT_GENERAL)

    # CHANGE 1: Exit Multiple updated to 10.5x (was 9.0x)
    style_label(ws, 45, 1, "Exit EV / EBITDA Multiple", indent=1)
    style_input(ws, 45, 2, 10.5, fmt=FMT_MULT)

    style_label(ws, 46, 1, "Exit Transaction Fees (% of TEV)", indent=1)
    style_input(ws, 46, 2, 0.01, fmt=FMT_PCT1)


# ════════════════════════════════════════════════════════════════════════════
# OPERATING MODEL SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_operating_model(ws):
    no_gridlines(ws)
    ws.sheet_properties.tabColor = "2E75B6"

    set_col_width(ws, 1, 36)
    set_col_width(ws, 2, 14)
    for c in range(3, 8):
        set_col_width(ws, c, 13)

    freeze(ws, "B4")

    # Row 1: Title
    c = ws.cell(row=1, column=1, value="Dick's Sporting Goods — Operating Model")
    c.font = _font(NAVY_FILL, bold=True, size=12)

    # Row 3: Column headers
    style_header_sub(ws, 3, 1, "")
    style_header_sub(ws, 3, 2, "LTM")
    for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
        style_header_sub(ws, 3, 3+i, yr)

    # ── INCOME STATEMENT / EBITDA BUILD ──────────────────────────────────────
    style_header_main(ws, 4, 1, "INCOME STATEMENT", col_span=7)

    # Row 5: Revenue
    style_label(ws, 5, 1, "Revenue", bold=True)
    style_link(ws, 5, 2, "='Assumptions'!B4", fmt=FMT_COMMA0)
    # Y1: =B5*(1+Assumptions!C37)
    col_letters = ['B','C','D','E','F','G']
    assm_growth_cols = ['C','D','E','F','G']  # C37..G37
    for i in range(5):  # Year 1..5
        col = 3 + i      # col C=3 .. G=7
        prev_col = col_letters[i]  # B, C, D, E, F
        growth_col = assm_growth_cols[i]
        formula = f"={prev_col}5*(1+'Assumptions'!{growth_col}37)"
        style_formula(ws, 5, col, formula, fmt=FMT_COMMA0)

    # Row 6: Revenue Growth %
    style_label(ws, 6, 1, "  Revenue Growth %")
    ws.cell(row=6, column=2).value = "LTM"
    ws.cell(row=6, column=2).font = _font(BLACK)
    ws.cell(row=6, column=2).alignment = _align("right")
    for i in range(5):
        col = 3 + i
        prev_col = col_letters[i]
        style_formula(ws, 6, col, f"={col_letters[i+1]}5/{prev_col}5-1", fmt=FMT_PCT1)

    # Row 8: EBITDA
    style_label(ws, 8, 1, "EBITDA", bold=True)
    style_link(ws, 8, 2, "='Assumptions'!B5", fmt=FMT_COMMA0)
    assm_margin_cols = ['C','D','E','F','G']
    for i in range(5):
        col = 3 + i
        margin_col = assm_margin_cols[i]
        rev_col = col_letters[i+1]
        formula = f"={rev_col}5*'Assumptions'!{margin_col}38"
        style_formula(ws, 8, col, formula, fmt=FMT_COMMA0)

    # Row 9: EBITDA Margin %
    style_label(ws, 9, 1, "  EBITDA Margin %")
    for i in range(6):
        col = 2 + i
        rev_col = col_letters[i]
        ebitda_col = col_letters[i]
        formula = f"={col_letters[i]}8/{col_letters[i]}5"
        style_formula(ws, 9, col, formula, fmt=FMT_PCT1)

    # Row 11: D&A
    style_label(ws, 11, 1, "  Depreciation & Amortization")
    assm_da_cols = ['C','D','E','F','G']
    for i in range(6):
        col = 2 + i
        da_col = assm_da_cols[i-1] if i > 0 else 'C'
        rev_col = col_letters[i]
        if i == 0:  # LTM col B
            formula = f"=B5*'Assumptions'!C39"
        else:
            da_assm_col = assm_da_cols[i-1]
            formula = f"={col_letters[i]}5*'Assumptions'!{assm_da_cols[i-1]}39"
        style_formula(ws, 11, col, formula, fmt=FMT_COMMA0)

    # Row 12: EBIT
    style_label(ws, 12, 1, "EBIT", bold=True)
    for i in range(6):
        col = 2 + i
        lc = col_letters[i]
        formula = f"={lc}8-{lc}11"
        style_formula(ws, 12, col, formula, fmt=FMT_COMMA0)

    # Row 13: EBIT Margin %
    style_label(ws, 13, 1, "  EBIT Margin %")
    for i in range(6):
        col = 2 + i
        lc = col_letters[i]
        formula = f"={lc}12/{lc}5"
        style_formula(ws, 13, col, formula, fmt=FMT_PCT1)

    # Row 15: Interest Expense — TLB
    # CHANGE 3: Link to Debt Schedule row 10 (was row 11 — TLB interest moved up one row)
    style_label(ws, 15, 1, "  Less: Interest Expense — TLB")
    ws.cell(row=15, column=2).value = 0
    ws.cell(row=15, column=2).font = _font(BLACK)
    ws.cell(row=15, column=2).number_format = FMT_COMMA0
    ws.cell(row=15, column=2).alignment = _align("right")
    # Y1-Y5: link from Debt Schedule row 10 (beginning balance interest, no circular)
    for i in range(5):
        col = 3 + i
        ds_col = col_letters[i+1]  # C..G
        style_link(ws, 15, col, f"='Debt Schedule'!{ds_col}10", fmt=FMT_COMMA0)

    # Row 16: Interest Expense — Revolver
    # CHANGE 3: Link to Debt Schedule row 17 (was row 18 — revolver section shifted up)
    style_label(ws, 16, 1, "  Less: Interest Expense — Revolver")
    ws.cell(row=16, column=2).value = 0
    ws.cell(row=16, column=2).font = _font(BLACK)
    ws.cell(row=16, column=2).number_format = FMT_COMMA0
    ws.cell(row=16, column=2).alignment = _align("right")
    for i in range(5):
        col = 3 + i
        ds_col = col_letters[i+1]
        style_link(ws, 16, col, f"='Debt Schedule'!{ds_col}17", fmt=FMT_COMMA0)

    # Row 17: Total Interest Expense
    style_label(ws, 17, 1, "Total Interest Expense", bold=True)
    ws.cell(row=17, column=2).value = 0
    ws.cell(row=17, column=2).font = _font(BLACK)
    ws.cell(row=17, column=2).number_format = FMT_COMMA0
    ws.cell(row=17, column=2).alignment = _align("right")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        style_formula(ws, 17, col, f"={lc}15+{lc}16", fmt=FMT_COMMA0)

    # Row 19: EBT
    style_label(ws, 19, 1, "EBT (Earnings Before Tax)", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"=MAX(0,{lc}12-{lc}17)"
        style_formula(ws, 19, col, formula, fmt=FMT_COMMA0)

    # Row 20: Income Tax Expense
    style_label(ws, 20, 1, "  Less: Income Tax Expense")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"=MAX(0,{lc}19)*'Assumptions'!B34"
        style_formula(ws, 20, col, formula, fmt=FMT_COMMA0)

    # Row 21: Net Income
    style_label(ws, 21, 1, "Net Income", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"={lc}19-{lc}20"
        style_total(ws, 21, col, formula, fmt=FMT_COMMA0)

    # ── FREE CASH FLOW BUILD ─────────────────────────────────────────────────
    style_header_main(ws, 23, 1, "FREE CASH FLOW", col_span=7)

    # Row 24: EBITDA
    style_label(ws, 24, 1, "EBITDA")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        style_formula(ws, 24, col, f"={lc}8", fmt=FMT_COMMA0)

    # Row 25: Less: Cash Taxes
    style_label(ws, 25, 1, "  Less: Cash Taxes")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"=-MAX(0,{lc}12-{lc}17)*'Assumptions'!B34"
        style_formula(ws, 25, col, formula, fmt=FMT_COMMA0)

    # Row 26: Less: Capital Expenditures
    style_label(ws, 26, 1, "  Less: Capital Expenditures")
    assm_capex_cols = ['C','D','E','F','G']
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        ac = assm_capex_cols[i]
        formula = f"=-{lc}5*'Assumptions'!{ac}40"
        style_formula(ws, 26, col, formula, fmt=FMT_COMMA0)

    # Row 27: Less: Change in NWC
    style_label(ws, 27, 1, "  Less: Change in Net Working Capital")
    assm_nwc_cols = ['C','D','E','F','G']
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        prev_lc = col_letters[i]
        ac = assm_nwc_cols[i]
        formula = f"=-({lc}5-{prev_lc}5)*'Assumptions'!{ac}41"
        style_formula(ws, 27, col, formula, fmt=FMT_COMMA0)

    # Row 28: Unlevered Free Cash Flow
    style_label(ws, 28, 1, "Unlevered Free Cash Flow", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"={lc}24+{lc}25+{lc}26+{lc}27"
        style_total(ws, 28, col, formula, fmt=FMT_COMMA0)

    # Row 30: Less: Total Interest Expense
    style_label(ws, 30, 1, "  Less: Total Interest Expense")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"=-{lc}17"
        style_formula(ws, 30, col, formula, fmt=FMT_COMMA0)

    # Row 31: Less: Mandatory TLB Amortization
    # NOTE: Debt Schedule row 6 is still mandatory amortization in the new structure
    style_label(ws, 31, 1, "  Less: Mandatory TLB Amortization")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        # Link from debt schedule row 6 (mandatory amortization, already negative)
        style_link(ws, 31, col, f"='Debt Schedule'!{lc}6", fmt=FMT_COMMA0)

    # Row 32: Levered Free Cash Flow
    # = Unlevered FCF (row 28) + Interest (row 30, already negated) + Mandatory Amort (row 31, already negative link)
    style_label(ws, 32, 1, "Levered Free Cash Flow", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"={lc}28+{lc}30+{lc}31"
        style_total(ws, 32, col, formula, fmt=FMT_COMMA0)


# ════════════════════════════════════════════════════════════════════════════
# DEBT SCHEDULE SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_debt_schedule(ws):
    no_gridlines(ws)
    ws.sheet_properties.tabColor = "2E75B6"

    set_col_width(ws, 1, 38)
    set_col_width(ws, 2, 14)
    for c in range(3, 8):
        set_col_width(ws, c, 13)

    freeze(ws, "B4")

    # Row 1: Title
    c = ws.cell(row=1, column=1, value="Dick's Sporting Goods — Debt Schedule")
    c.font = _font(NAVY_FILL, bold=True, size=12)

    # Row 3: Column headers
    style_header_sub(ws, 3, 1, "")
    for i, yr in enumerate(["Year 1","Year 2","Year 3","Year 4","Year 5"]):
        style_header_sub(ws, 3, 3+i, yr)

    col_letters = ['B','C','D','E','F','G']

    # ── TERM LOAN B ───────────────────────────────────────────────────────────
    # CHANGE 2: Restructured TLB section with cash sweep + beginning-of-period interest
    style_header_main(ws, 4, 1, "TERM LOAN B", col_span=7)

    # Row 5: Beginning Balance
    style_label(ws, 5, 1, "  Beginning Balance")
    # Y1: link from Assumptions TLB
    style_link(ws, 5, 3, "='Assumptions'!B17", fmt=FMT_COMMA0)
    # Y2-Y5: = prior year ending balance (row 8)
    for i in range(1, 5):  # Y2..Y5
        col = 3 + i  # D=4, E=5, F=6, G=7
        prev_col = col_letters[i]  # C, D, E, F
        style_formula(ws, 5, col, f"={prev_col}8", fmt=FMT_COMMA0)

    # Row 6: Mandatory Amortization (negative — cash out)
    style_label(ws, 6, 1, "  Less: Mandatory Amortization")
    for i in range(5):
        col = 3 + i
        formula = "=-'Assumptions'!B17*'Assumptions'!B30"
        style_formula(ws, 6, col, formula, fmt=FMT_COMMA0)

    # Row 7: Cash Sweep (negative — uses beginning-of-period Levered FCF, no circular)
    # Pulls Levered FCF from Operating Model row 32. Interest in row 32 uses beginning
    # balance (row 5 x row 9), which is known, so NO circular reference.
    style_label(ws, 7, 1, "  Less: Cash Sweep")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]  # C, D, E, F, G
        formula = f"=-MAX(0,'Operating Model'!{lc}32)"
        style_formula(ws, 7, col, formula, fmt=FMT_COMMA0)

    # Row 8: Ending Balance
    style_label(ws, 8, 1, "  Ending Balance", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"=MAX(0,{lc}5+{lc}6+{lc}7)"
        style_total(ws, 8, col, formula, fmt=FMT_COMMA0)

    # Row 9: Interest Rate (green link — no Average Balance row)
    style_label(ws, 9, 1, "  Interest Rate")
    for i in range(5):
        col = 3 + i
        style_link(ws, 9, col, "='Assumptions'!B29", fmt=FMT_PCT2)

    # Row 10: Interest Expense — TLB (uses BEGINNING balance x rate — breaks circular)
    # This is intentional: beginning balance is known, so no circular reference.
    style_label(ws, 10, 1, "  Interest Expense — TLB", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        # Use row 5 (beginning balance) x row 9 (rate) — NOT average
        formula = f"={lc}5*{lc}9"
        style_total(ws, 10, col, formula, fmt=FMT_COMMA0)

    # ── REVOLVING CREDIT FACILITY ─────────────────────────────────────────────
    # CHANGE 2: Section shifted up by 1 row (now starts at row 12 instead of 13)
    style_header_main(ws, 12, 1, "REVOLVING CREDIT FACILITY", col_span=7)

    # Row 13: Beginning Balance
    style_label(ws, 13, 1, "  Beginning Balance")
    for i in range(5):
        col = 3 + i
        style_formula(ws, 13, col, "=0", fmt=FMT_COMMA0)

    # Row 14: Draw / (Repayment)
    style_label(ws, 14, 1, "  Draw / (Repayment)")
    for i in range(5):
        col = 3 + i
        style_formula(ws, 14, col, "=0", fmt=FMT_COMMA0)

    # Row 15: Ending Balance
    style_label(ws, 15, 1, "  Ending Balance", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        style_total(ws, 15, col, f"={lc}13+{lc}14", fmt=FMT_COMMA0)

    # Row 16: Interest Rate (All-In)
    style_label(ws, 16, 1, "  Interest Rate (All-In)")
    for i in range(5):
        col = 3 + i
        style_link(ws, 16, col, "='Assumptions'!B32", fmt=FMT_PCT2)

    # Row 17: Interest Expense — Revolver (commitment fee on undrawn capacity)
    style_label(ws, 17, 1, "  Interest Expense — Revolver (Commitment Fee)", bold=True)
    for i in range(5):
        col = 3 + i
        # Commitment fee on full capacity since revolver stays undrawn
        formula = "='Assumptions'!B18*'Assumptions'!B33"
        style_total(ws, 17, col, formula, fmt=FMT_COMMA0)

    # ── TOTAL DEBT SUMMARY ────────────────────────────────────────────────────
    # CHANGE 2: Summary section shifts to row 19
    style_header_main(ws, 19, 1, "TOTAL DEBT SUMMARY", col_span=7)

    # Row 20: Total Debt — Beginning
    style_label(ws, 20, 1, "  Total Debt — Beginning")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"={lc}5+{lc}13"
        style_formula(ws, 20, col, formula, fmt=FMT_COMMA0)

    # Row 21: Total Debt — Ending
    style_label(ws, 21, 1, "  Total Debt — Ending", bold=True)
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"={lc}8+{lc}15"
        style_total(ws, 21, col, formula, fmt=FMT_COMMA0)

    # Row 22: Total Net Debt at Exit
    style_label(ws, 22, 1, "  Total Net Debt at Exit")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        formula = f"={lc}21"
        style_formula(ws, 22, col, formula, fmt=FMT_COMMA0)

    # Row 23: Leverage at Exit
    style_label(ws, 23, 1, "  Leverage at Exit (x EBITDA)")
    for i in range(5):
        col = 3 + i
        lc = col_letters[i+1]
        # EBITDA for that year from Operating Model
        formula = f"={lc}21/'Operating Model'!{lc}8"
        style_formula(ws, 23, col, formula, fmt=FMT_MULT)


# ════════════════════════════════════════════════════════════════════════════
# RETURNS SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_returns(ws):
    no_gridlines(ws)
    ws.sheet_properties.tabColor = "375623"

    set_col_width(ws, 1, 40)
    set_col_width(ws, 2, 18)

    # Row 1: Title
    c = ws.cell(row=1, column=1, value="Dick's Sporting Goods — Returns Analysis")
    c.font = _font(NAVY_FILL, bold=True, size=12)

    # ── EXIT VALUATION ────────────────────────────────────────────────────────
    style_header_main(ws, 3, 1, "EXIT VALUATION", col_span=2)

    style_label(ws, 4, 1, "Exit Year EBITDA ($M)", indent=1)
    # Y5 EBITDA from Operating Model col G (Year 5)
    style_link(ws, 4, 2, "='Operating Model'!G8", fmt=FMT_COMMA0)

    style_label(ws, 5, 1, "Exit EV / EBITDA Multiple", indent=1)
    style_link(ws, 5, 2, "='Assumptions'!B45", fmt=FMT_MULT)

    style_label(ws, 6, 1, "Exit Enterprise Value ($M)", indent=1)
    style_formula(ws, 6, 2, "=B4*B5", fmt=FMT_COMMA0)

    style_label(ws, 7, 1, "Less: Net Debt at Exit ($M)", indent=1)
    # CHANGE 4: TLB ending balance Year 5 (col G = row 8 in Debt Schedule) — stays correct
    style_link(ws, 7, 2, "='Debt Schedule'!G8", fmt=FMT_COMMA0)

    style_label(ws, 8, 1, "Less: Exit Transaction Fees ($M)", indent=1)
    style_formula(ws, 8, 2, "=B6*'Assumptions'!B46", fmt=FMT_COMMA0)

    style_label(ws, 9, 1, "Equity Value to Sponsor ($M)", bold=True, indent=1)
    style_total(ws, 9, 2, "=B6-B7-B8", fmt=FMT_COMMA0)

    # ── RETURNS ANALYSIS ──────────────────────────────────────────────────────
    style_header_main(ws, 11, 1, "RETURNS ANALYSIS", col_span=2)

    style_label(ws, 12, 1, "Initial Equity Investment ($M)", indent=1)
    style_link(ws, 12, 2, "='Assumptions'!B21", fmt=FMT_COMMA0)

    style_label(ws, 13, 1, "Equity Proceeds ($M)", indent=1)
    style_formula(ws, 13, 2, "=B9", fmt=FMT_COMMA0)

    style_label(ws, 15, 1, "Multiple of Invested Capital (MOIC)", bold=True, indent=1)
    style_total(ws, 15, 2, "=B13/B12", fmt='0.0"x"')

    style_label(ws, 16, 1, "Hold Period (Years)", indent=1)
    style_link(ws, 16, 2, "='Assumptions'!B44", fmt=FMT_GENERAL)

    style_label(ws, 17, 1, "IRR (Approximate)", bold=True, indent=1)
    style_total(ws, 17, 2, "=(B13/B12)^(1/B16)-1", fmt=FMT_PCT1)

    style_label(ws, 18, 1, "Cash-on-Cash Return (MOIC)", indent=1)
    style_formula(ws, 18, 2, "=B15", fmt='0.0"x"')


# ════════════════════════════════════════════════════════════════════════════
# SENSITIVITY SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_sensitivity(ws):
    no_gridlines(ws)
    ws.sheet_properties.tabColor = "C55A11"

    set_col_width(ws, 1, 18)
    for c in range(2, 8):
        set_col_width(ws, c, 13)

    # Row 1: Title
    c = ws.cell(row=1, column=1, value="Dick's Sporting Goods — Sensitivity Analysis")
    c.font = _font(NAVY_FILL, bold=True, size=12)

    entry_multiples = [7.0, 8.0, 9.0, 10.0, 11.0]
    exit_multiples  = [7.0, 8.0, 9.0, 10.0, 11.0]

    def build_table(start_row, title, metric):
        """Build a sensitivity table. metric='IRR' or 'MOIC'."""
        # Section header
        style_header_main(ws, start_row, 1, title, col_span=6)

        # Sub-header row
        style_header_sub(ws, start_row+1, 1, "Entry \\ Exit", h_align="center")
        for j, em in enumerate(exit_multiples):
            style_header_sub(ws, start_row+1, 2+j, f"{em:.1f}x")

        base_entry_col = None
        base_exit_col  = None

        for i, entry_m in enumerate(entry_multiples):
            row = start_row + 2 + i
            # Row header
            c = ws.cell(row=row, column=1, value=f"{entry_m:.1f}x")
            c.fill = _fill(NAVY2_FILL)
            c.font = _font(WHITE, bold=True)
            c.alignment = _align("center")

            for j, exit_m in enumerate(exit_multiples):
                col = 2 + j

                # Build the formula using literals for the entry/exit multiples
                # and references for everything else
                entry_equity_f      = f"({entry_m}*'Assumptions'!B5-'Assumptions'!B9-{entry_m}*'Assumptions'!B5*'Assumptions'!B11)"
                exit_equity_f       = f"({exit_m}*'Operating Model'!G8-'Debt Schedule'!G8-{exit_m}*'Operating Model'!G8*'Assumptions'!B46)"

                if metric == 'IRR':
                    formula = (f"=IFERROR(({exit_equity_f}/{entry_equity_f})"
                               f"^(1/'Assumptions'!B44)-1,\"N/A\")")
                    fmt = FMT_PCT1
                else:  # MOIC
                    formula = f"=IFERROR({exit_equity_f}/{entry_equity_f},\"N/A\")"
                    fmt = '0.0"x"'

                cell = ws.cell(row=row, column=col, value=formula)
                cell.font = _font(BLACK)
                cell.alignment = _align("center")
                cell.number_format = fmt

                # Track base case (9.0x / 9.0x) = i=2, j=2
                if entry_m == 9.0 and exit_m == 9.0:
                    base_entry_col = col
                    base_row = row

        # Bold border on base case
        if base_entry_col is not None:
            bc = ws.cell(row=base_row, column=base_entry_col)
            thick = Side(style="medium")
            bc.border = Border(left=thick, right=thick, top=thick, bottom=thick)

        # Conditional formatting — color scale over the data range
        data_start = f"{get_column_letter(2)}{start_row+2}"
        data_end   = f"{get_column_letter(6)}{start_row+6}"
        ws.conditional_formatting.add(
            f"{data_start}:{data_end}",
            ColorScaleRule(
                start_type='min', start_color='F8696B',
                mid_type='percentile', mid_value=50, mid_color='FFEB84',
                end_type='max', end_color='63BE7B'
            )
        )

    build_table(start_row=3,  title="TABLE 1: IRR SENSITIVITY — Entry Multiple vs. Exit Multiple",  metric='IRR')
    build_table(start_row=13, title="TABLE 2: MOIC SENSITIVITY — Entry Multiple vs. Exit Multiple", metric='MOIC')

    # Labels
    for start_row, title in [(3, "IRR"), (13, "MOIC")]:
        c = ws.cell(row=start_row+8, column=1,
                    value=f"Rows = Entry EV/EBITDA Multiple | Columns = Exit EV/EBITDA Multiple")
        c.font = _font("7F7F7F", italic=True, size=9)


# ════════════════════════════════════════════════════════════════════════════
# COVER SHEET
# ════════════════════════════════════════════════════════════════════════════
def build_cover(ws):
    no_gridlines(ws)
    ws.sheet_properties.tabColor = "1F3864"

    # Column layout: A=margin, B=left-label, C=left-value, D=spacer,
    #                E=right-label, F=right-value, G=margin
    set_col_width(ws, 1, 2)   # A — left margin
    set_col_width(ws, 2, 28)  # B — label / content
    set_col_width(ws, 3, 14)  # C — value
    set_col_width(ws, 4, 4)   # D — spacer
    set_col_width(ws, 5, 28)  # E — right label
    set_col_width(ws, 6, 14)  # F — right value
    set_col_width(ws, 7, 2)   # G — right margin

    def merge_full(row, height=None):
        """Merge B through F for full-width rows."""
        ws.merge_cells(start_row=row, start_column=2,
                       end_row=row, end_column=6)
        if height:
            ws.row_dimensions[row].height = height

    # ── TITLE BLOCK ───────────────────────────────────────────────────────────
    merge_full(3, height=32)
    c = ws.cell(row=3, column=2, value="LEVERAGED BUYOUT ANALYSIS")
    c.font = Font(name="Calibri", size=20, bold=True, color=NAVY_FILL)
    c.alignment = _align("center")

    merge_full(4, height=4)   # spacer

    merge_full(5, height=20)
    c = ws.cell(row=5, column=2, value="Dick's Sporting Goods, Inc.  (NYSE: DKS)")
    c.font = _font(NAVY_FILL, bold=True, size=13)
    c.alignment = _align("center")

    merge_full(6, height=16)
    c = ws.cell(row=6, column=2,
                value="Consumer Retail  \u2022  Big-Box Sporting Goods  \u2022  Transaction Date: 2026")
    c.font = _font("595959", size=10)
    c.alignment = _align("center")

    # thin rule under header
    merge_full(7, height=6)
    ws.cell(row=7, column=2).border = Border(
        bottom=Side(style="medium", color=NAVY_FILL))

    # ── TRANSACTION SUMMARY (2-column stats grid) ─────────────────────────────
    # Header spans full width
    ws.merge_cells(start_row=8, start_column=2, end_row=8, end_column=6)
    style_header_main(ws, 8, 2, "TRANSACTION SUMMARY", col_span=5)
    ws.row_dimensions[8].height = 16

    # 3 rows × 2 stats each
    stat_pairs = [
        ("Transaction Enterprise Value ($M)", "='Assumptions'!B8",  FMT_COMMA0,
         "Entry EV / LTM EBITDA Multiple",    "='Assumptions'!B7",  FMT_MULT),
        ("Total Leverage at Close (x)",        "='Assumptions'!B23", FMT_MULT,
         "Sponsor Equity ($M)",               "='Assumptions'!B21", FMT_COMMA0),
        ("Projected IRR — Year 5",             "=Returns!B17",       FMT_PCT1,
         "Projected MOIC — Year 5",           "=Returns!B15",       '0.0"x"'),
    ]
    for i, (l_lbl, l_frm, l_fmt, r_lbl, r_frm, r_fmt) in enumerate(stat_pairs):
        r = 9 + i
        ws.row_dimensions[r].height = 18
        # Left label (col B)
        cl = ws.cell(row=r, column=2, value=l_lbl)
        cl.fill = _fill(GRAY_FILL); cl.font = _font(NAVY_FILL, bold=True)
        cl.alignment = _align("left")
        # Left value (col C)
        cv = ws.cell(row=r, column=3, value=l_frm)
        cv.fill = _fill(GRAY_FILL); cv.font = _font(DARK_BLUE_FONT, bold=True)
        cv.alignment = _align("right"); cv.number_format = l_fmt
        # Spacer (col D)
        ws.cell(row=r, column=4).fill = _fill("FFFFFF")
        # Right label (col E)
        el = ws.cell(row=r, column=5, value=r_lbl)
        el.fill = _fill(GRAY_FILL); el.font = _font(NAVY_FILL, bold=True)
        el.alignment = _align("left")
        # Right value (col F)
        ev = ws.cell(row=r, column=6, value=r_frm)
        ev.fill = _fill(GRAY_FILL); ev.font = _font(DARK_BLUE_FONT, bold=True)
        ev.alignment = _align("right"); ev.number_format = r_fmt

    # ── DEAL RATIONALE ────────────────────────────────────────────────────────
    ws.merge_cells(start_row=13, start_column=2, end_row=13, end_column=6)
    style_header_main(ws, 13, 2, "DEAL RATIONALE", col_span=5)
    ws.row_dimensions[13].height = 16

    narrative = (
        "DKS is the last major national big-box sporting goods retailer standing — Sports Authority "
        "filed for bankruptcy in 2016 and Modell\u2019s followed in 2020. Entry at 9.0x LTM EBITDA "
        "acquires a market-share-accreting platform with a clear operational thesis: accelerating "
        "private label penetration from ~15% toward 20%+ of revenue mix drives EBITDA margins from "
        "13% to 16.5% over the hold period. Aggressive free cash flow sweeps against the Term Loan B "
        "deleverage the business from 5.5x at entry to approximately 2.4x at exit \u2014 a "
        "transformation in earnings quality that justifies a re-rating to 10.5x on exit and "
        "delivers 3.8x MOIC and 30%+ IRR over a 5-year hold."
    )
    ws.merge_cells(start_row=14, start_column=2, end_row=14, end_column=6)
    c = ws.cell(row=14, column=2, value=narrative)
    c.font = Font(name="Calibri", size=9.5, italic=True, color="1F1F1F")
    c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[14].height = 60

    # ── VALUE CREATION DRIVERS ────────────────────────────────────────────────
    ws.merge_cells(start_row=16, start_column=2, end_row=16, end_column=6)
    style_header_main(ws, 16, 2, "VALUE CREATION DRIVERS", col_span=5)
    ws.row_dimensions[16].height = 16

    drivers = [
        ("MARKET CONSOLIDATION",
         "Sole surviving national platform in a $50B+ addressable market; no credible "
         "large-format competitor to rebuild — market share gains are structural and durable."),
        ("PRIVATE LABEL EXPANSION",
         "DSG and Alpine Design carry 500\u2013800bps EBITDA margin premium vs. national brands; "
         "growing mix from ~15% to 20%+ drives ~350bps of margin expansion across the hold period."),
        ("EXPERIENTIAL DIFFERENTIATION",
         "House of Sport flagship format (climbing walls, batting cages, TrackMan simulators) "
         "drives higher AUR and traffic that e-commerce cannot replicate."),
        ("CASH FLOW DELEVERAGING",
         "$3.4B in cumulative TLB sweeps reduce net debt from $9.4B to $5.5B; compounding equity "
         "value creation through debt paydown, EBITDA growth, and exit multiple re-rating."),
    ]
    for i, (title, desc) in enumerate(drivers):
        r = 17 + i
        ws.row_dimensions[r].height = 28
        # Category title (col B) — all-caps bold navy
        ct = ws.cell(row=r, column=2, value=title)
        ct.font = Font(name="Calibri", size=9, bold=True, color=NAVY_FILL)
        ct.alignment = _align("left")
        # Description (cols C-F merged)
        ws.merge_cells(start_row=r, start_column=3,
                       end_row=r, end_column=6)
        cd = ws.cell(row=r, column=3, value=desc)
        cd.font = Font(name="Calibri", size=9, color="1F1F1F")
        cd.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ── CONFIDENTIALITY ───────────────────────────────────────────────────────
    merge_full(22, height=14)
    c = ws.cell(row=22, column=2,
                value="STRICTLY CONFIDENTIAL \u2014 FOR DISCUSSION PURPOSES ONLY")
    c.font = Font(name="Calibri", size=8.5, italic=True, color="9E9E9E")
    c.alignment = _align("center")


# ════════════════════════════════════════════════════════════════════════════
# MAIN BUILD FUNCTION
# ════════════════════════════════════════════════════════════════════════════
def build_model():
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets in order
    ws_cover   = wb.create_sheet("Cover")
    ws_assm    = wb.create_sheet("Assumptions")
    ws_op      = wb.create_sheet("Operating Model")
    ws_debt    = wb.create_sheet("Debt Schedule")
    ws_returns = wb.create_sheet("Returns")
    ws_sens    = wb.create_sheet("Sensitivity")

    print("Building Cover sheet...")
    build_cover(ws_cover)

    print("Building Assumptions sheet...")
    build_assumptions(ws_assm)

    print("Building Operating Model sheet...")
    build_operating_model(ws_op)

    print("Building Debt Schedule sheet...")
    build_debt_schedule(ws_debt)

    print("Building Returns sheet...")
    build_returns(ws_returns)

    print("Building Sensitivity sheet...")
    build_sensitivity(ws_sens)

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print(f"Done! File saved to: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_model()
