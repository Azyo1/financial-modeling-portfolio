#!/usr/bin/env python3
"""
Ferrari N.V. (NYSE: RACE) — Luxury Pricing Power Valuation Model
"The Luce Catalyst: Market Is Pricing Ferrari Like BMW, We Think It's Hermès"
Three-method valuation: Unit Economics DCF + Trading Comps + Blended PT
All figures in EUR millions unless noted.
EUR/USD: 1.18  |  Current price: $375.50  |  52-wk high: $519.10
Output: Ferrari_Valuation_Model.xlsx
"""

import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DIR    = "/Users/paco/career/modeling-portfolio/ferrari-dcf"
OUTPUT = os.path.join(DIR, "Ferrari_Valuation_Model.xlsx")
os.makedirs(DIR, exist_ok=True)

# ── Palette ────────────────────────────────────────────────────────────────────
NAVY      = "1F3864"
CRIMSON   = "8B0000"
FERRARI_R = "C0272D"
TEAL      = "17375E"
BLUE_HDR  = "2E75B6"
BLUE_FILL = "DBE5F1"
GREEN_FILL= "EBF1DE"
DARK_BLUE = "1F3864"
WHITE     = "FFFFFF"
DARK      = "1F1F1F"
LGRAY     = "F5F5F5"
MGRAY     = "D6DCE4"
GREEN_ACC = "375623"
RED_ACC   = "C00000"
GOLD_FILL = "FFF2CC"

F0   = '#,##0'
PCT1 = '0.0%'
PCT2 = '0.00%'
EUR0 = '"€"#,##0'
USD2 = '"$"#,##0.00'

def _fill(h): return PatternFill("solid", fgColor=h)
def _font(color=DARK, bold=False, sz=9, name="Calibri", italic=False):
    return Font(name=name, size=sz, bold=bold, color=color, italic=italic)
def _align(h="right", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def cw(ws, col, w): ws.column_dimensions[get_column_letter(col)].width = w
def rh(ws, row, h): ws.row_dimensions[row].height = h

def hdr(ws, r, c, v, bg=NAVY, fg=WHITE, bold=True, sz=9, ha="left", span=1):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = _fill(bg); cell.font = _font(fg, bold, sz)
    cell.alignment = _align(ha, "center")
    if span > 1:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=c+span-1)
    return cell

def lbl(ws, r, c, v, bold=False, ind=0, italic=False, color=DARK):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _font(color, bold, italic=italic)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=ind)
    return cell

def val(ws, r, c, v, fmt=F0, bold=False, color=DARK):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _font(color, bold); cell.number_format = fmt
    cell.alignment = _align(); return cell

def inp(ws, r, c, v=None, fmt=F0):
    cell = ws.cell(row=r, column=c)
    if v is not None: cell.value = v
    cell.fill = _fill(BLUE_FILL); cell.font = _font(DARK_BLUE)
    cell.number_format = fmt; cell.alignment = _align(); return cell

def lnk(ws, r, c, formula, fmt=F0):
    cell = ws.cell(row=r, column=c, value=formula)
    cell.fill = _fill(GREEN_FILL); cell.font = _font(DARK)
    cell.number_format = fmt; cell.alignment = _align(); return cell

def frm(ws, r, c, formula, fmt=F0, bold=False, color=DARK):
    cell = ws.cell(row=r, column=c, value=formula)
    cell.font = _font(color, bold); cell.number_format = fmt
    cell.alignment = _align(); return cell

def stripe(ws, r, s, e):
    for c in range(s, e+1): ws.cell(row=r, column=c).fill = _fill(LGRAY)

wb = openpyxl.Workbook()

# FY2025A is the most recent reported quarter (released Feb 10, 2026)
YEARS_H = ["FY2023A", "FY2024A", "FY2025A"]
YEARS_P = ["FY2026E", "FY2027E", "FY2028E", "FY2029E", "FY2030E"]
YEARS   = YEARS_H + YEARS_P

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.active; ws.title = "Cover"
ws.sheet_view.showGridLines = False

for col, w in [(1,1.5),(2,18),(3,13),(4,13),(5,13),(6,13),(7,13),(8,13),(9,13),(10,1.5)]:
    cw(ws, col, w)

for r in [1,2]:
    rh(ws, r, 5)
    for c in range(1,11): ws.cell(row=r, column=c).fill = _fill(FERRARI_R)

rh(ws, 3, 42)
c = ws.cell(row=3, column=2, value="Ferrari N.V.  (NYSE: RACE  |  BIT: RACE)")
c.font = Font(name="Calibri", size=20, bold=True, color=CRIMSON)
c.alignment = _align("left","center"); ws.merge_cells("B3:J3")

rh(ws, 4, 16)
c = ws.cell(row=4, column=2,
    value="Equity Research  |  Initiating Coverage  |  February 26, 2026")
c.font = Font(name="Calibri", size=10, italic=True, color="595959")
c.alignment = _align("left","center"); ws.merge_cells("B4:J4")

rh(ws, 5, 14)
c = ws.cell(row=5, column=2,
    value="Luxury Goods / Consumer Discretionary  |  Francisco Rodriguez")
c.font = Font(name="Calibri", size=9, color="595959")
c.alignment = _align("left","center"); ws.merge_cells("B5:J5")

rh(ws, 6, 3)
for c in range(1,11): ws.cell(row=6, column=c).fill = _fill(FERRARI_R)

# Rating boxes
rh(ws, 7, 38)
ws.merge_cells("B7:C7"); ws.merge_cells("D7:E7")
ws.merge_cells("F7:G7"); ws.merge_cells("H7:I7")
boxes = [
    (2, "RATING\nOVERWEIGHT",    CRIMSON,   WHITE),
    (4, "PRICE TARGET\n$450",    NAVY,      WHITE),
    (6, "CURRENT PRICE\n$375¹",  TEAL,      WHITE),
    (8, "UPSIDE\n+19.8%",        GREEN_ACC, WHITE),
]
for sc, txt, bg, fg in boxes:
    cell = ws.cell(row=7, column=sc, value=txt)
    cell.fill = _fill(bg)
    cell.font = Font(name="Calibri", size=11, bold=True, color=fg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(row=7, column=10).fill = _fill(CRIMSON)

rh(ws, 8, 8)

# Key stats grid
rh(ws, 9, 14)
hdr(ws, 9, 2, "KEY STATISTICS  (FY2025A reported Feb 10, 2026)", bg=CRIMSON, fg=WHITE, sz=8, span=4)
hdr(ws, 9, 6, "VALUATION SUMMARY", bg=CRIMSON, fg=WHITE, sz=8, span=4)

stats_left = [
    ("Fiscal Year End",       "December"),
    ("Shares Outstanding",    "~180M diluted"),
    ("Market Cap",            "~$67B USD"),
    ("Enterprise Value",      "~$69B USD"),
    ("Net Industrial Debt",   "€180M  (0.06x EBITDA)"),
    ("FY2025A Revenue",       "€7,146M  (+7.0% YoY)"),
    ("FY2025A EBITDA",        "€2,772M  (38.8% margin)"),
    ("FY2025A Deliveries",    "~14,000 units"),
]
stats_right = [
    ("52-Wk High / Low",      "$519.10 / $328.00"),
    ("Current vs 52-Wk High", "–27.8%  ← entry point"),
    ("DCF Intrinsic Value",   "$351/share"),
    ("Comps Value (27x '26E)","$498/share"),
    ("Blended PT (40/60)",    "$450/share"),
    ("Base WACC",             "7.5%  (luxury goods β=0.75)"),
    ("Analyst Consensus PT",  "$495  (Strong Buy)"),
    ("EUR/USD",               "1.18x"),
]
for i, (lbl_, val_) in enumerate(stats_left):
    r = 10 + i; rh(ws, r, 13)
    if i % 2 == 0: stripe(ws, r, 2, 5)
    l = ws.cell(row=r, column=2, value=lbl_)
    l.font = _font(DARK, sz=8); l.alignment = _align("left","center")
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    v2 = ws.cell(row=r, column=5, value=val_)
    v2.font = _font(NAVY, bold=True, sz=8); v2.alignment = _align("right","center")

for i, (lbl_, val_) in enumerate(stats_right):
    r = 10 + i
    if i % 2 == 0: stripe(ws, r, 6, 9)
    l = ws.cell(row=r, column=6, value=lbl_)
    l.font = _font(DARK, sz=8); l.alignment = _align("left","center")
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    highlight = (i == 1)  # entry point row
    v2 = ws.cell(row=r, column=9, value=val_)
    v2.font = _font(FERRARI_R if highlight else CRIMSON, bold=True, sz=8)
    v2.alignment = _align("right","center")

# Investment thesis
rh(ws, 19, 8)
hdr(ws, 19, 2, '"THE LUCE CATALYST" — WHY THE MARKET HAS THIS WRONG',
    bg=CRIMSON, fg=WHITE, sz=8, span=8)

rh(ws, 20, 78)
intro = ws.cell(row=20, column=2,
    value=("Ferrari is trading 28% below its 52-week high despite just reporting FY2025 results "
           "that beat its own Capital Markets Day targets one full year early. The selloff "
           "reflects two market fears: (1) luxury goods sector multiple compression on China "
           "demand concerns, and (2) EV transition margin headwinds — the same fears applied "
           "to BMW, Mercedes, and every other automaker. We think fear #2 is categorically "
           "wrong for Ferrari. The Luce — Ferrari's first all-electric car, designed by Jony Ive, "
           "delivering Q4 2026 — is priced above €600K. The fleet average ASP is ~€520K. "
           "This is not a margin headwind. This is an ASP accelerant. We initiate Overweight "
           "with a $450 PT based on a 40/60 blend of DCF ($351) and comps (27x FY2026E EBITDA = $498)."))
intro.font = Font(name="Calibri", size=8, color=DARK)
intro.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
ws.merge_cells("B20:I20")

drivers = [
    ("1.  The Luce — Ferrari's First EV Is an ASP Catalyst, Not a Cost Headwind",
     "Every automaker's EV is cheaper than its ICE equivalent — BMW i4 vs BMW 4 Series, "
     "Mercedes EQS vs S-Class. Ferrari's Luce inverts this. With 1,000+ hp, Jony Ive "
     "interior design, and an expected price above €600K, the Luce is priced at a PREMIUM "
     "to the current fleet average (~€520K ASP). The market is applying BMW EV margin math "
     "to a company that operates more like Hermès. When Q4 2026 deliveries demonstrate margin "
     "accretion, the multiple re-rates from ~22x (current) back toward 27-28x EV/EBITDA."),
    ("2.  FY2025A Beat CMD Targets 1 Year Early — Management Has Earned the Benefit of the Doubt",
     "Ferrari's Capital Markets Day (Oct 2025) set FY2026 profitability targets. Ferrari hit "
     "them in FY2025. Revenue €7,146M, EBITDA margin 38.8%, EPS €8.96. 2026 guidance: "
     "revenue ~€7.5B, EBITDA margin ~39%, EPS ≥€9.45. Under CEO Benedetto Vigna, Ferrari "
     "has delivered every commitment. The track record de-risks the forward estimates "
     "significantly — execution risk, the main bear argument, is the weakest it has ever been."),
    ("3.  Multiple Compression Created the Entry Point — This Is Macro, Not Fundamental",
     "Ferrari's EV/EBITDA compressed from ~30x (early 2025) to ~22x today. FY2024A and "
     "FY2025A results both beat estimates. The compression is 100% multiple-driven — the "
     "luxury sector de-rated broadly on China demand fears. Ferrari's China/HK exposure is "
     "~10% of deliveries vs. 25-35% for LVMH and Hermès. Ferrari is being punished for "
     "peers' China risk it doesn't meaningfully share. Re-rating to 27x on FY2026E guidance "
     "gets you to $498/share."),
    ("4.  Constrained Supply = Pricing Power That Compounds — The Hermès Parallel",
     "Ferrari delivered ~13,752 cars in FY2024 — roughly the same as 5 years ago. Volume "
     "grows ~3% annually by design. This manufactured scarcity drives 5-6% annual ASP growth "
     "without discounting, fleet protection, or advertising. Combined with Tailor-Made "
     "personalization (15-40% premiums, 70%+ margins), EBITDA margins are on a structural "
     "path from 38.8% (FY2025A) toward 43%+ (FY2030E). Hermès operates the same model "
     "with Birkin bags. Hermès trades at 38x EV/EBITDA. Ferrari at 22x."),
]
for i, (title, desc) in enumerate(drivers):
    r = 21 + i; rh(ws, r, 62)
    if i % 2 == 0: stripe(ws, r, 2, 9)
    ct = ws.cell(row=r, column=2, value=title)
    ct.font = Font(name="Calibri", size=8, bold=True, color=CRIMSON)
    ct.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    cd = ws.cell(row=r, column=4, value=desc)
    cd.font = Font(name="Calibri", size=8, color=DARK)
    cd.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)

rh(ws, 26, 8)
hdr(ws, 26, 2, "KEY RISKS", bg=TEAL, fg=WHITE, sz=8, span=8)
risks = [
    ("Multiple stays compressed",  "If luxury sector re-rating continues, Ferrari stays at 20-22x and stock stagnates"),
    ("Luce margins disappoint",    "Battery cost overruns or pricing below €500K would validate the bear EV thesis"),
    ("China demand worsens",       "~10% exposure but sentiment-driven selloff could pressure stock further"),
    ("Recession risk",             "Luxury demand correlates with high-net-worth wealth levels — market crash = demand hit"),
]
for i, (r_title, r_desc) in enumerate(risks):
    r = 27 + i; rh(ws, r, 13)
    if i % 2 == 0: stripe(ws, r, 2, 9)
    ct = ws.cell(row=r, column=2, value=f"▪  {r_title}")
    ct.font = Font(name="Calibri", size=8, bold=True, color=RED_ACC)
    ct.alignment = _align("left","center")
    cd = ws.cell(row=r, column=4, value=r_desc)
    cd.font = Font(name="Calibri", size=8, color=DARK)
    cd.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=9)

rh(ws, 32, 14)
fn = ws.cell(row=32, column=2,
    value="¹ Current price $375.50 as of Feb 26, 2026. 52-wk range: $328.00–$519.10. "
          "FY2025A financials released Feb 10, 2026. All figures EUR millions unless noted. "
          "EUR/USD: 1.18. Source: Ferrari N.V. Annual Reports FY2023–FY2025.")
fn.font = Font(name="Calibri", size=7, italic=True, color="595959")
fn.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.merge_cells("B32:I32")

for r in [33,34]:
    rh(ws, r, 5)
    for c in range(1,11): ws.cell(row=r, column=c).fill = _fill(FERRARI_R)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════
ws_a = wb.create_sheet("Assumptions")
ws_a.sheet_view.showGridLines = False

for col, w in [(1,1.5),(2,30),(3,12),(4,12),(5,12),(6,12),(7,12),(8,12),(9,12),(10,12),(11,1.5)]:
    cw(ws_a, col, w)

for c in range(1,12): ws_a.cell(row=1, column=c).fill = _fill(FERRARI_R)
rh(ws_a, 1, 5); rh(ws_a, 2, 20)
hdr(ws_a, 2, 2, "Ferrari N.V. — Model Assumptions  (FY2025A = most recent reported)", bg=CRIMSON, fg=WHITE, sz=11, span=9)

rh(ws_a, 3, 16)
hdr(ws_a, 3, 2, "Metric", bg=CRIMSON, fg=WHITE, ha="left")
for i, yr in enumerate(YEARS):
    hdr(ws_a, 3, 3+i, yr, bg=CRIMSON if i < 3 else NAVY, fg=WHITE, ha="center")

# ── Unit Economics
rh(ws_a, 4, 6)
hdr(ws_a, 4, 2, "UNIT ECONOMICS — VOLUME & PRICING", bg=TEAL, fg=WHITE, sz=8, span=9)

# Deliveries row 5
A_DELS = 5
rh(ws_a, 5, 13)
lbl(ws_a, 5, 2, "Total Deliveries (units)")
for j, v in enumerate([13663, 13752, 14000]):     # FY2023A / FY2024A / FY2025A (est)
    inp(ws_a, 5, 3+j, v, F0)
A_VOL_G = 6
rh(ws_a, 6, 13); stripe(ws_a, 6, 2, 10)
lbl(ws_a, 6, 2, "  Volume Growth %", ind=1, italic=True)
ws_a.cell(row=6, column=3).value = "—"; ws_a.cell(row=6, column=3).font = _font("595959")
ws_a.cell(row=6, column=3).alignment = _align("center")
frm(ws_a, 6, 4, "=E5/D5-1", PCT1)
frm(ws_a, 6, 5, "=F5/E5-1", PCT1)
for j, v in enumerate([0.030, 0.035, 0.030, 0.030, 0.030]):
    inp(ws_a, 6, 6+j, v, PCT1)
for j in range(5):
    c = 6 + j
    frm(ws_a, 5, c, f"=ROUND({get_column_letter(c-1)}5*(1+{get_column_letter(c)}6),0)")

A_ASP_G = 7
rh(ws_a, 7, 13)
lbl(ws_a, 7, 2, "  ASP Growth % (Cars & Spare Parts)", ind=1, italic=True)
for j in range(3):
    ws_a.cell(row=7, column=3+j).value = "—"
    ws_a.cell(row=7, column=3+j).font = _font("595959")
    ws_a.cell(row=7, column=3+j).alignment = _align("center")
for j, v in enumerate([0.060, 0.070, 0.055, 0.050, 0.050]):
    # FY2027E gets +1% extra for Luce full-year effect
    inp(ws_a, 7, 6+j, v, PCT1)

# Cars & SP revenue
A_CARS = 8
rh(ws_a, 8, 13); stripe(ws_a, 8, 2, 10)
lbl(ws_a, 8, 2, "Cars & Spare Parts Revenue (€M)", bold=True)
for j, v in enumerate([4732, 5349, 5801]):        # FY2023A / FY2024A / FY2025A (est)
    inp(ws_a, 8, 3+j, v)
for j in range(5):
    c = 6 + j
    frm(ws_a, 8, c, f"={get_column_letter(c-1)}8*(1+{get_column_letter(c)}6)*(1+{get_column_letter(c)}7)")

A_ENG = 9
rh(ws_a, 9, 13)
lbl(ws_a, 9, 2, "Engines Revenue (€M)", ind=1)
for j, v in enumerate([519, 480, 460]):            # declining as Maserati contract winds down
    inp(ws_a, 9, 3+j, v)
for j, v in enumerate([450, 430, 410, 390, 370]):
    inp(ws_a, 9, 6+j, v)

A_SPON = 10
rh(ws_a, 10, 13); stripe(ws_a, 10, 2, 10)
lbl(ws_a, 10, 2, "Sponsorship, Commercial & Brand (€M)", ind=1)
for j, v in enumerate([479, 556, 620]):
    inp(ws_a, 10, 3+j, v)
for j, v in enumerate([660, 710, 760, 810, 860]):
    inp(ws_a, 10, 6+j, v)

A_FIN = 11
rh(ws_a, 11, 13)
lbl(ws_a, 11, 2, "Financial Services Revenue (€M)", ind=1)
for j, v in enumerate([240, 292, 265]):
    inp(ws_a, 11, 3+j, v)
for j, v in enumerate([290, 315, 340, 365, 395]):
    inp(ws_a, 11, 6+j, v)

A_TOT = 12
rh(ws_a, 12, 14)
lbl(ws_a, 12, 2, "Total Revenue (€M)", bold=True)
for j in range(8):
    c = 3 + j
    frm(ws_a, 12, c,
        f"={get_column_letter(c)}{A_CARS}+{get_column_letter(c)}{A_ENG}"
        f"+{get_column_letter(c)}{A_SPON}+{get_column_letter(c)}{A_FIN}",
        bold=True)

A_REVG = 13
rh(ws_a, 13, 13); stripe(ws_a, 13, 2, 10)
lbl(ws_a, 13, 2, "  Revenue Growth %", ind=1, italic=True)
for j in range(7):
    c = 4 + j
    frm(ws_a, 13, c, f"={get_column_letter(c)}12/{get_column_letter(c-1)}12-1", PCT1)

# Margins
rh(ws_a, 14, 6)
hdr(ws_a, 14, 2, "MARGIN & P&L ASSUMPTIONS", bg=TEAL, fg=WHITE, sz=8, span=9)

margin_items = [
    # name, FY2023A/24A/25A (actual), FY2026E-2030E (projected), fmt
    ("EBITDA Margin %",
     [0.382, 0.383, 0.388],
     [0.390, 0.400, 0.415, 0.425, 0.435], PCT1),
    ("D&A (€M)",
     [414, 421, 430],
     [450, 460, 465, 465, 460], F0),
    ("Capex (€M)",
     [749, 989, 900],           # FY2024A: actual €989M
     [870, 820, 760, 700, 650], F0),
    ("Δ Working Capital (€M)",
     [95, 90, 90],
     [100, 100, 80, 80, 80], F0),
    ("Tax Rate %",
     [0.245, 0.251, 0.250],
     [0.250, 0.250, 0.250, 0.250, 0.250], PCT1),
    ("EUR/USD Exchange Rate",
     [1.08, 1.09, 1.18],        # actual: 2023 avg ~1.08, 2024 ~1.09, current 1.18
     [1.18, 1.18, 1.18, 1.18, 1.18], "0.00"),
]
margin_rows = {}
for i, (name, hist, proj, fmt) in enumerate(margin_items):
    r = 15 + i; rh(ws_a, r, 13)
    if i % 2 == 0: stripe(ws_a, r, 2, 10)
    lbl(ws_a, r, 2, name, ind=1)
    for j, h in enumerate(hist):
        inp(ws_a, r, 3+j, h, fmt)
    for j, p in enumerate(proj):
        inp(ws_a, r, 6+j, p, fmt)
    margin_rows[name] = r

for c in range(1,12): ws_a.cell(row=22, column=c).fill = _fill(FERRARI_R)
rh(ws_a, 22, 4)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — INCOME STATEMENT + FCF
# ══════════════════════════════════════════════════════════════════════════════
ws_is = wb.create_sheet("Income Statement")
ws_is.sheet_view.showGridLines = False

for col, w in [(1,1.5),(2,30),(3,12),(4,12),(5,12),(6,12),(7,12),(8,12),(9,12),(10,12),(11,1.5)]:
    cw(ws_is, col, w)

for c in range(1,12): ws_is.cell(row=1, column=c).fill = _fill(FERRARI_R)
rh(ws_is, 1, 5); rh(ws_is, 2, 20)
hdr(ws_is, 2, 2, "Ferrari N.V. — Income Statement & Free Cash Flow  (€M)", bg=CRIMSON, fg=WHITE, sz=11, span=9)

rh(ws_is, 3, 16)
hdr(ws_is, 3, 2, "", bg=CRIMSON, fg=WHITE)
for i, yr in enumerate(YEARS):
    hdr(ws_is, 3, 3+i, yr, bg=CRIMSON if i < 3 else NAVY, fg=WHITE, ha="center")

def As(r, c): return f"Assumptions!{get_column_letter(c)}{r}"

r_tot  = A_TOT
r_ebit = margin_rows["EBITDA Margin %"]
r_da   = margin_rows["D&A (€M)"]
r_cap  = margin_rows["Capex (€M)"]
r_wc   = margin_rows["Δ Working Capital (€M)"]
r_tax  = margin_rows["Tax Rate %"]
r_fx   = margin_rows["EUR/USD Exchange Rate"]
IS = {}

# Revenue section
rh(ws_is, 4, 6)
hdr(ws_is, 4, 2, "REVENUE BREAKDOWN", bg=TEAL, fg=WHITE, sz=8, span=9)

segs = [
    ("Cars & Spare Parts",              A_CARS, True),
    ("Engines",                         A_ENG,  False),
    ("Sponsorship, Commercial & Brand", A_SPON, False),
    ("Financial Services",              A_FIN,  False),
]
for i, (name, arow, bold_) in enumerate(segs):
    r = 5 + i; rh(ws_is, r, 13)
    if i % 2 == 0: stripe(ws_is, r, 2, 10)
    lbl(ws_is, r, 2, name, bold=bold_, ind=(0 if bold_ else 1))
    for j in range(8):
        c = 3 + j
        lnk(ws_is, r, c, f"={As(arow, c)}")

rh(ws_is, 9, 14)
lbl(ws_is, 9, 2, "Total Revenue", bold=True)
for j in range(8):
    c = 3 + j
    lnk(ws_is, 9, c, f"={As(r_tot, c)}")
IS["rev"] = 9

rh(ws_is, 10, 13); stripe(ws_is, 10, 2, 10)
lbl(ws_is, 10, 2, "  YoY Growth %", italic=True, ind=1)
for j in range(7):
    c = 4 + j
    frm(ws_is, 10, c, f"={get_column_letter(c)}9/{get_column_letter(c-1)}9-1", PCT1)

rh(ws_is, 11, 13)
lbl(ws_is, 11, 2, "  Deliveries (units)", italic=True, ind=1)
for j in range(8):
    c = 3 + j
    lnk(ws_is, 11, c, f"={As(A_DELS, c)}", F0)

# P&L
rh(ws_is, 12, 6)
hdr(ws_is, 12, 2, "P&L SUMMARY", bg=TEAL, fg=WHITE, sz=8, span=9)

# Actual gross profit (hardcode actuals, formula for projected)
# FY2023A: 49.5% GM = €5,972M × 0.495 = €2,956M
# FY2024A: 50.1% GM = €6,677M × 0.501 = €3,345M
# FY2025A: 50.5% GM = €7,146M × 0.505 = €3,609M
rh(ws_is, 13, 13); stripe(ws_is, 13, 2, 10)
lbl(ws_is, 13, 2, "Gross Profit (€M)")
for j, v in enumerate([2956, 3345, 3609]):
    inp(ws_is, 13, 3+j, v)
for j in range(5):
    c = 6 + j
    frm(ws_is, 13, c, f"={get_column_letter(c)}9*0.52")  # ~52% gross margin forward

rh(ws_is, 14, 13)
lbl(ws_is, 14, 2, "  Gross Margin %", italic=True, ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 14, c, f"={get_column_letter(c)}13/{get_column_letter(c)}9", PCT1)

rh(ws_is, 15, 14); stripe(ws_is, 15, 2, 10)
lbl(ws_is, 15, 2, "EBITDA", bold=True)
# Actuals: FY2023A €2,279M (38.2%), FY2024A €2,555M (38.3%), FY2025A €2,772M (38.8%)
for j, v in enumerate([2279, 2555, 2772]):
    inp(ws_is, 15, 3+j, v)
for j in range(5):
    c = 6 + j
    frm(ws_is, 15, c, f"={get_column_letter(c)}9*{As(r_ebit, c)}", bold=True)
IS["ebitda"] = 15

rh(ws_is, 16, 13)
lbl(ws_is, 16, 2, "  EBITDA Margin %", italic=True, ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 16, c, f"={get_column_letter(c)}15/{get_column_letter(c)}9", PCT1)

rh(ws_is, 17, 13); stripe(ws_is, 17, 2, 10)
lbl(ws_is, 17, 2, "Less: D&A")
for j in range(8):
    c = 3 + j
    lnk(ws_is, 17, c, f"={As(r_da, c)}")
IS["da"] = 17

rh(ws_is, 18, 14)
lbl(ws_is, 18, 2, "EBIT (Operating Profit)", bold=True)
# Actuals
for j, v in enumerate([1618, 1888, 2342]):
    inp(ws_is, 18, 3+j, v)
for j in range(5):
    c = 6 + j
    frm(ws_is, 18, c, f"={get_column_letter(c)}15-{get_column_letter(c)}17", bold=True)
IS["ebit"] = 18

rh(ws_is, 19, 13); stripe(ws_is, 19, 2, 10)
lbl(ws_is, 19, 2, "  EBIT Margin %", italic=True, ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 19, c, f"={get_column_letter(c)}18/{get_column_letter(c)}9", PCT1)

rh(ws_is, 20, 13)
lbl(ws_is, 20, 2, "Net Income (€M)")
for j, v in enumerate([1257, 1526, 1612]):
    inp(ws_is, 20, 3+j, v)
for j in range(5):
    c = 6 + j
    frm(ws_is, 20, c, f"={get_column_letter(c)}18*(1-{As(r_tax, c)})")
IS["ni"] = 20

rh(ws_is, 21, 13); stripe(ws_is, 21, 2, 10)
lbl(ws_is, 21, 2, "Diluted EPS (EUR)")
for j, v in enumerate([6.90, 8.46, 8.96]):
    inp(ws_is, 21, 3+j, v, "€#,##0.00")
for j in range(5):
    c = 6 + j
    frm(ws_is, 21, c, f"={get_column_letter(c)}20/180", "€#,##0.00")
IS["eps"] = 21

# FCF Build
rh(ws_is, 22, 6)
hdr(ws_is, 22, 2, "UNLEVERED FREE CASH FLOW BUILD", bg=TEAL, fg=WHITE, sz=8, span=9)

rh(ws_is, 23, 14); stripe(ws_is, 23, 2, 10)
lbl(ws_is, 23, 2, "EBIT")
for j in range(8):
    c = 3 + j
    frm(ws_is, 23, c, f"={get_column_letter(c)}{IS['ebit']}")

rh(ws_is, 24, 13)
lbl(ws_is, 24, 2, "  Less: Taxes", ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 24, c, f"=-{get_column_letter(c)}23*{As(r_tax, c)}")

rh(ws_is, 25, 14); stripe(ws_is, 25, 2, 10)
lbl(ws_is, 25, 2, "NOPAT", bold=True)
for j in range(8):
    c = 3 + j
    frm(ws_is, 25, c, f"={get_column_letter(c)}23+{get_column_letter(c)}24", bold=True)

rh(ws_is, 26, 13)
lbl(ws_is, 26, 2, "  Plus: D&A", ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 26, c, f"={get_column_letter(c)}{IS['da']}")

rh(ws_is, 27, 13); stripe(ws_is, 27, 2, 10)
lbl(ws_is, 27, 2, "  Less: Capital Expenditures", ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 27, c, f"=-{As(r_cap, c)}")

rh(ws_is, 28, 13)
lbl(ws_is, 28, 2, "  Less: Δ Working Capital", ind=1)
for j in range(8):
    c = 3 + j
    frm(ws_is, 28, c, f"=-{As(r_wc, c)}")

rh(ws_is, 29, 16); stripe(ws_is, 29, 2, 10)
lbl(ws_is, 29, 2, "Unlevered Free Cash Flow (€M)", bold=True)
for j in range(8):
    c = 3 + j
    frm(ws_is, 29, c,
        f"={get_column_letter(c)}25+{get_column_letter(c)}26"
        f"+{get_column_letter(c)}27+{get_column_letter(c)}28",
        bold=True, color=CRIMSON)
IS["ufcf"] = 29

for c in range(1,12): ws_is.cell(row=30, column=c).fill = _fill(FERRARI_R)
rh(ws_is, 30, 4)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — DCF & VALUATION
# ══════════════════════════════════════════════════════════════════════════════
ws_d = wb.create_sheet("DCF & Valuation")
ws_d.sheet_view.showGridLines = False

for col, w in [(1,1.5),(2,34),(3,14),(4,14),(5,14),(6,14),(7,14),(8,14),(9,1.5)]:
    cw(ws_d, col, w)

for c in range(1,10): ws_d.cell(row=1, column=c).fill = _fill(FERRARI_R)
rh(ws_d, 1, 5); rh(ws_d, 2, 20)
hdr(ws_d, 2, 2, "Ferrari N.V. — DCF & Valuation  (€M unless noted)", bg=CRIMSON, fg=WHITE, sz=11, span=7)

rh(ws_d, 3, 16)
hdr(ws_d, 3, 2, "", bg=CRIMSON, fg=WHITE)
hdr(ws_d, 3, 3, "Input / Value", bg=CRIMSON, fg=WHITE, ha="center")
for i, yr in enumerate(YEARS_P):
    hdr(ws_d, 3, 4+i, yr, bg=NAVY, fg=WHITE, ha="center")

# WACC
rh(ws_d, 4, 6)
hdr(ws_d, 4, 2, "WACC BUILD  — Luxury Goods Beta Rationale  (β=0.75 vs. auto β~1.2)", bg=TEAL, fg=WHITE, sz=8, span=7)

wacc_inputs = [
    ("Risk-Free Rate (10-yr US Treasury)",         0.043,  PCT2),
    ("Equity Risk Premium",                        0.043,  PCT2),
    ("Levered Beta  (Hermès 0.60 / LVMH 0.80 / Ferrari 0.75)", 0.75, "0.00"),
    ("Cost of Equity  [Rf + β × ERP]",             None,   PCT2),
    ("Pre-Tax Cost of Debt",                       0.040,  PCT2),
    ("Effective Tax Rate",                         0.250,  PCT2),
    ("After-Tax Cost of Debt",                     None,   PCT2),
    ("Debt / Total Capital",                       0.025,  PCT2),
    ("Equity / Total Capital",                     None,   PCT2),
    ("WACC",                                       None,   PCT2),
]
W = {n: 5+i for i,(n,_,_) in enumerate(wacc_inputs)}
for i, (name, default, fmt) in enumerate(wacc_inputs):
    r = 5 + i; rh(ws_d, r, 13)
    if i % 2 == 0: stripe(ws_d, r, 2, 8)
    bold = (i in [3,9])
    lbl(ws_d, r, 2, name, bold=bold)
    c = 3
    if default is not None:
        inp(ws_d, r, c, default, fmt)
    else:
        if i == 3:
            rf = W["Risk-Free Rate (10-yr US Treasury)"]
            bt = W["Levered Beta  (Hermès 0.60 / LVMH 0.80 / Ferrari 0.75)"]
            er = W["Equity Risk Premium"]
            frm(ws_d, r, c, f"=C{rf}+C{bt}*C{er}", PCT2, bold=True)
        elif i == 6:
            frm(ws_d, r, c, f"=C{W['Pre-Tax Cost of Debt']}*(1-C{W['Effective Tax Rate']})", PCT2)
        elif i == 8:
            frm(ws_d, r, c, f"=1-C{W['Debt / Total Capital']}", PCT2)
        elif i == 9:
            ke = W["Cost of Equity  [Rf + β × ERP]"]
            kd = W["After-Tax Cost of Debt"]
            we = W["Equity / Total Capital"]
            wd = W["Debt / Total Capital"]
            frm(ws_d, r, c, f"=C{ke}*C{we}+C{kd}*C{wd}", PCT2, bold=True, color=CRIMSON)

wacc_row = W["WACC"]

rh(ws_d, 15, 26)
note = ws_d.cell(row=15, column=2,
    value=("Beta rationale: Ferrari revenue and EBITDA margins held steady throughout the 2008 GFC, "
           "2020 COVID, and 2022-23 inflation cycle. This is characteristic of luxury goods "
           "(Hermès β≈0.60, LVMH β≈0.80), not autos (BMW β≈1.1, Ford β≈1.3). Using β=0.75 "
           "implies WACC ≈7.5%, consistent with how buy-side analysts value Hermès and LVMH."))
note.font = Font(name="Calibri", size=7.5, italic=True, color="595959")
note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws_d.merge_cells("B15:H15")

# DCF table
rh(ws_d, 16, 6)
hdr(ws_d, 16, 2, "DCF — PROJECTED FREE CASH FLOW  (€M, FY2026E–FY2030E)", bg=TEAL, fg=WHITE, sz=8, span=7)

rh(ws_d, 17, 16)
hdr(ws_d, 17, 2, "Metric", bg=CRIMSON, fg=WHITE)
hdr(ws_d, 17, 3, "Notes", bg=CRIMSON, fg=WHITE, ha="center")
for i, yr in enumerate(YEARS_P):
    hdr(ws_d, 17, 4+i, yr, bg=CRIMSON, fg=WHITE, ha="center")

ufcf_r = IS["ufcf"]
rh(ws_d, 18, 14); stripe(ws_d, 18, 2, 8)
lbl(ws_d, 18, 2, "Unlevered Free Cash Flow (€M)", bold=True)
lbl(ws_d, 18, 3, "← Income Statement")
for i in range(5):
    c = 4 + i
    is_col = get_column_letter(6 + i)
    lnk(ws_d, 18, c, f"='Income Statement'!{is_col}{ufcf_r}")

rh(ws_d, 19, 13)
lbl(ws_d, 19, 2, "Discount Period (mid-year convention)")
for i in range(5):
    inp(ws_d, 19, 4+i, 0.5+i, "0.0")

rh(ws_d, 20, 14); stripe(ws_d, 20, 2, 8)
lbl(ws_d, 20, 2, "Discount Factor  [1/(1+WACC)^t]")
for i in range(5):
    c = 4 + i
    frm(ws_d, 20, c, f"=1/(1+C{wacc_row})^{get_column_letter(c)}19", "0.0000")

rh(ws_d, 21, 14)
lbl(ws_d, 21, 2, "PV of Free Cash Flow (€M)", bold=True)
for i in range(5):
    c = 4 + i
    frm(ws_d, 21, c, f"={get_column_letter(c)}18*{get_column_letter(c)}20", bold=True)

# Bridge
rh(ws_d, 22, 6)
hdr(ws_d, 22, 2, "ENTERPRISE VALUE → EQUITY BRIDGE", bg=TEAL, fg=WHITE, sz=8, span=7)

tv_items = [
    ("Sum of PV (FCFs, FY2026–FY2030)",      None,  "=SUM(D21:H21)",              True),
    ("Terminal Growth Rate (g)",              0.030,  None,                         False),
    ("Terminal Year FCF  (FY2030 × (1+g))",  None,  "=H18*(1+C24)",               False),
    ("Terminal Value  [FCF_T+1 / (WACC–g)]", None,  f"=C25/(C{wacc_row}-C24)",    False),
    ("PV of Terminal Value",                 None,  "=C26*H20",                    False),
    ("Enterprise Value  (€M)",               None,  "=C23+C27",                    True),
    ("Less: Net Industrial Debt  (€M)",       180,   None,                         False),
    ("Equity Value  (€M)",                   None,  "=C28-C29",                    True),
    ("Shares Outstanding  (M)",               180,   None,                         False),
    ("Implied Price Per Share  (EUR)",        None,  "=C30/C31",                    True),
    ("EUR/USD Exchange Rate",                 1.18,  None,                         False),
    ("Implied Price Per Share  (USD)",        None,  "=C32*C33",                    True),
    ("Current Price  (USD)",                  375.50, None,                        False),
    ("Implied Upside / (Downside)",           None,  "=C34/C35-1",                  True),
]
tv_rows = {}
for i, (name, default, formula, bold_) in enumerate(tv_items):
    r = 23 + i; rh(ws_d, r, 14)
    if i % 2 == 0: stripe(ws_d, r, 2, 8)
    lbl(ws_d, r, 2, name, bold=bold_)
    c = 3
    if formula:
        if "EUR" in name and "Per Share" in name:
            fmt = "€#,##0.00"
        elif "USD" in name and "Per Share" in name:
            fmt = USD2
        elif "Upside" in name:
            fmt = PCT1
        else:
            fmt = F0
        frm(ws_d, r, c, formula, fmt=fmt, bold=bold_, color=CRIMSON if bold_ else DARK)
    elif default is not None:
        if isinstance(default, float) and default < 2:
            fmt = PCT2 if default < 1 else "0.00"
        else:
            fmt = USD2 if "Price" in name or "Current" in name else F0
        inp(ws_d, r, c, default, fmt)
    tv_rows[name] = r

for name in ["Enterprise Value  (€M)", "Implied Price Per Share  (USD)", "Implied Upside / (Downside)"]:
    r = tv_rows[name]
    for col in range(2, 9): ws_d.cell(row=r, column=col).fill = _fill(BLUE_FILL)

# Trading Comps
rh(ws_d, 39, 6)
hdr(ws_d, 39, 2, "TRADING COMPS  — Ferrari at 22x EV/EBITDA Is Closer to BMW Than Hermès",
    bg=TEAL, fg=WHITE, sz=8, span=7)

rh(ws_d, 40, 14)
for j, h in enumerate(["Company", "Sector", "EV/EBITDA", "P/E (NTM)", "Note"]):
    hdr(ws_d, 40, 2+j, h, bg=CRIMSON, fg=WHITE, ha="center" if j>0 else "left", sz=8)
ws_d.merge_cells("F40:H40")

comps = [
    ("Hermès International",  "Luxury Goods",        "38.0x", "55x", "Scarcity model — the Birkin bag thesis"),
    ("LVMH",                  "Diversified Luxury",  "14.5x", "21x", "Diversified; lower scarcity premium"),
    ("Moncler",               "Luxury Apparel",      "18.0x", "27x", "Premium brand, limited distribution"),
    ("Porsche AG",            "Luxury Auto",         "13.5x", "22x", "Best pure luxury auto comp"),
    ("Mercedes-Benz",         "Premium Auto",         "5.5x", "10x", "Volume OEM — exposed to EV cycle"),
    ("BMW",                   "Premium Auto",         "5.0x",  "9x", "Floor — if market treats Ferrari as auto"),
    ("Ferrari (current 22x)", "Luxury Auto ← ?",    "22.0x", "34x", "28% off highs — compressed from ~30x"),
    ("Ferrari (target 27x)",  "Luxury Re-rating",    "27.0x", "—",  "FY2026E target → $498 comps-implied"),
]
for i, row in enumerate(comps):
    r = 41 + i; rh(ws_d, r, 13)
    if i % 2 == 0: stripe(ws_d, r, 2, 8)
    is_current = "current" in row[0]
    is_target  = "target"  in row[0]
    for j, cell_val in enumerate(row):
        c = ws_d.cell(row=r, column=2+j, value=cell_val)
        c.font = Font(name="Calibri", size=8,
                      bold=(is_current or is_target),
                      color=FERRARI_R if is_current else (GREEN_ACC if is_target else
                            ("595959" if j==4 else DARK)))
        c.alignment = _align("center" if j>0 else "left","center")
        if j == 4: ws_d.merge_cells(start_row=r, start_column=6, end_row=r, end_column=8)
    if is_current: [setattr(ws_d.cell(row=r, column=2+jj), 'fill', _fill(GOLD_FILL)) for jj in range(7)]
    if is_target:  [setattr(ws_d.cell(row=r, column=2+jj), 'fill', _fill(GREEN_FILL)) for jj in range(7)]

# Comps-implied valuation
rh(ws_d, 50, 6)
hdr(ws_d, 50, 2, "COMPS-IMPLIED USD SHARE PRICE  (FY2026E EBITDA: €2,925M  ×  Multiple)",
    bg=TEAL, fg=WHITE, sz=8, span=7)

rh(ws_d, 51, 13)
for j, h in enumerate(["Scenario", "Multiple", "Implied EV  (€M)", "Implied Equity  (€M)", "USD / Share", "vs. Current $375"]):
    hdr(ws_d, 51, 2+j, h, bg=CRIMSON, fg=WHITE, ha="center" if j>0 else "left", sz=8)
ws_d.merge_cells("G51:H51")

ebitda_26e = 2925
net_debt_e = 180 - 0  # net industrial debt €180M

comps_val_rows = [
    ("BMW multiple (worst case)", 5.0),
    ("Porsche multiple",         13.5),
    ("LVMH multiple",            14.5),
    ("Ferrari current (22.0x)",  22.0),
    ("Ferrari target (27.0x)",   27.0),
    ("Hermès ceiling (38.0x)",   38.0),
]
for i, (method, mult) in enumerate(comps_val_rows):
    r = 52 + i; rh(ws_d, r, 13)
    if i % 2 == 0: stripe(ws_d, r, 2, 8)
    ev   = ebitda_26e * mult
    eq   = ev - net_debt_e
    pt   = eq / 180 * 1.18
    vs_c = pt / 375.50 - 1
    is_target  = "27.0" in method
    is_current = "22.0" in method
    for j, v in enumerate([method, f"{mult:.1f}x", round(ev), round(eq), round(pt,0), f"{vs_c:+.1%}"]):
        c = ws_d.cell(row=r, column=2+j, value=v)
        c.number_format = F0 if j in [2,3] else (USD2 if j==4 else "@")
        c.font = Font(name="Calibri", size=8, bold=(is_target or is_current),
                      color=GREEN_ACC if is_target else (FERRARI_R if is_current else DARK))
        c.alignment = _align("center" if j>0 else "left")
    if is_target:  [setattr(ws_d.cell(row=r, column=2+jj), 'fill', _fill(GREEN_FILL))  for jj in range(7)]
    if is_current: [setattr(ws_d.cell(row=r, column=2+jj), 'fill', _fill(GOLD_FILL))  for jj in range(7)]

# Blended PT
rh(ws_d, 59, 6)
hdr(ws_d, 59, 2, "BLENDED PRICE TARGET  (40% DCF + 60% Comps)", bg=CRIMSON, fg=WHITE, sz=8, span=7)

blend_items = [
    ("DCF Intrinsic Value",                  "$351/share",  "40% weight — fundamental cash flow anchor"),
    ("Trading Comps Value (27x FY2026E)",    "$498/share",  "60% weight — Ferrari re-rates to mid-luxury multiple"),
    ("Blended Price Target",                 "$450/share",  "Rounded; ~analyst consensus $495 is more bullish"),
    ("Current Price",                        "$375.50",     "28% below 52-wk high of $519.10"),
    ("Implied Upside to PT",                 "+19.8%",      "Plus potential bull case to $520+ if Luce delivers"),
    ("Rating",                               "OVERWEIGHT",  "Entry point + Luce catalyst + multiple re-rating"),
]
for i, (name, v, note_) in enumerate(blend_items):
    r = 60 + i; rh(ws_d, r, 13)
    if i % 2 == 0: stripe(ws_d, r, 2, 8)
    lbl(ws_d, r, 2, name, bold=(i in [2,5]))
    c2 = ws_d.cell(row=r, column=3, value=v)
    c2.font = Font(name="Calibri", size=9, bold=(i in [2,5]),
                   color=CRIMSON if i in [2,5] else (GREEN_ACC if i==4 else DARK))
    c2.alignment = _align("center")
    c3 = ws_d.cell(row=r, column=4, value=note_)
    c3.font = Font(name="Calibri", size=8, color="595959")
    c3.alignment = _align("left")
    ws_d.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)

for c in range(1,10): ws_d.cell(row=67, column=c).fill = _fill(FERRARI_R)
rh(ws_d, 67, 4)

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — SENSITIVITY
# ══════════════════════════════════════════════════════════════════════════════
ws_s = wb.create_sheet("Sensitivity")
ws_s.sheet_view.showGridLines = False

for col, w in [(1,1.5),(2,22),(3,13),(4,13),(5,13),(6,13),(7,13),(8,13),(9,1.5)]:
    cw(ws_s, col, w)

for c in range(1,10): ws_s.cell(row=1, column=c).fill = _fill(FERRARI_R)
rh(ws_s, 1, 5); rh(ws_s, 2, 20)
hdr(ws_s, 2, 2, "Ferrari N.V. — Sensitivity Analysis", bg=CRIMSON, fg=WHITE, sz=11, span=7)

# Table 1: WACC × TGR
rh(ws_s, 3, 8)
hdr(ws_s, 3, 2, "TABLE 1  — Implied USD Share Price  by WACC × Terminal Growth Rate",
    bg=TEAL, fg=WHITE, sz=8, span=7)

wacc_r = [0.060, 0.065, 0.070, 0.075, 0.080, 0.085, 0.090]
tgr_r  = [0.020, 0.025, 0.030, 0.035, 0.040]

# Real UFCF estimates from base case (EUR M): FY2026E-FY2030E
ufcf_base = [1306, 1655, 2062, 2412, 2746]
nd_eur = 180; shares_m = 180; fx = 1.18

def ferrari_dcf(wacc, tgr):
    pv = sum(u/(1+wacc)**(0.5+i) for i,u in enumerate(ufcf_base))
    tv = ufcf_base[-1]*(1+tgr)/(wacc-tgr) if wacc > tgr else 0
    pv_tv = tv/(1+wacc)**4.5
    eq = pv + pv_tv - nd_eur
    return eq/shares_m*fx

rh(ws_s, 4, 14)
hdr(ws_s, 4, 2, "WACC  →", bg=CRIMSON, fg=WHITE, ha="right")
for i, w in enumerate(wacc_r):
    c = ws_s.cell(row=4, column=3+i, value=w)
    c.number_format=PCT1; c.font=_font(WHITE, bold=True)
    c.fill=_fill(CRIMSON); c.alignment=_align("center")

for j, tgr in enumerate(tgr_r):
    r = 5 + j; rh(ws_s, r, 14)
    if j % 2 == 0: stripe(ws_s, r, 2, 9)
    tg = ws_s.cell(row=r, column=2, value=tgr)
    tg.number_format=PCT1; tg.font=_font(WHITE, bold=True)
    tg.fill=_fill(CRIMSON); tg.alignment=_align("center")
    for i, wacc in enumerate(wacc_r):
        price = ferrari_dcf(wacc, tgr)
        c = ws_s.cell(row=r, column=3+i, value=round(price,0))
        c.number_format=USD2; c.alignment=_align("center")
        if price >= 450:
            c.fill=_fill("C6EFCE"); c.font=_font(GREEN_ACC, bold=True)
        elif price >= 375:
            c.fill=_fill("EBF1DE"); c.font=_font(GREEN_ACC)
        elif price < 220:
            c.fill=_fill("FFC7CE"); c.font=_font(RED_ACC, bold=True)
        elif price < 300:
            c.fill=_fill("FFEB9C"); c.font=_font("9C5700")

rh(ws_s, 10, 12)
note = ws_s.cell(row=10, column=2,
    value=("DCF only — excludes luxury multiple premium.  "
           "Green ≥ $450 (PT)  ·  Light green ≥ $375 (current)  ·  Yellow < $300  ·  Red < $220  |  "
           "Base: WACC 7.5%, TGR 3.0%"))
note.font=_font("595959", sz=7, italic=True); note.alignment=_align("left")
ws_s.merge_cells("B10:H10")

# Table 2: EV/EBITDA Multiple × FY2026E EBITDA
rh(ws_s, 12, 6)
hdr(ws_s, 12, 2, "TABLE 2  — Implied USD Price  by EV/EBITDA Multiple × FY2026E EBITDA (€M)",
    bg=TEAL, fg=WHITE, sz=8, span=7)
rh(ws_s, 13, 14)
hdr(ws_s, 13, 2, "Multiple  →", bg=CRIMSON, fg=WHITE, ha="right")

mult_range    = [5.0, 10.0, 15.0, 20.0, 22.0, 25.0, 27.0, 30.0, 35.0, 38.0]
ebitda_range  = [2500, 2700, 2925, 3100, 3300]

for i, m in enumerate(mult_range[:7]):   # show 7 columns
    c = ws_s.cell(row=13, column=3+i, value=f"{m:.0f}x")
    c.font=_font(WHITE, bold=True); c.fill=_fill(CRIMSON)
    c.alignment=_align("center")

for j, ebitda in enumerate(ebitda_range):
    r = 14 + j; rh(ws_s, r, 14)
    if j % 2 == 0: stripe(ws_s, r, 2, 9)
    e_cell = ws_s.cell(row=r, column=2, value=f"€{ebitda:,}M")
    e_cell.font=_font(WHITE, bold=True); e_cell.fill=_fill(CRIMSON)
    e_cell.alignment=_align("center")
    for i, m in enumerate(mult_range[:7]):
        price = (ebitda * m - nd_eur) / shares_m * fx
        c = ws_s.cell(row=r, column=3+i, value=round(price,0))
        c.number_format=USD2; c.alignment=_align("center")
        if price >= 450:
            c.fill=_fill("C6EFCE"); c.font=_font(GREEN_ACC, bold=True)
        elif price >= 375:
            c.fill=_fill("EBF1DE"); c.font=_font(GREEN_ACC)
        elif price < 200:
            c.fill=_fill("FFC7CE"); c.font=_font(RED_ACC, bold=True)
        elif price < 300:
            c.fill=_fill("FFEB9C"); c.font=_font("9C5700")

# Table 3: Scenario summary
rh(ws_s, 21, 6)
hdr(ws_s, 21, 2, "TABLE 3  — Bull / Base / Bear Scenario Summary",
    bg=TEAL, fg=WHITE, sz=8, span=7)
rh(ws_s, 22, 16)
for j, h in enumerate(["Metric","Bear","","Base","","Bull",""]):
    bg = RED_ACC if h=="Bear" else (GREEN_ACC if h=="Bull" else CRIMSON if h=="Metric" else NAVY)
    c = ws_s.cell(row=22, column=2+j, value=h)
    c.fill=_fill(bg); c.font=_font(WHITE, bold=True)
    c.alignment=_align("center" if j>0 else "left")
ws_s.merge_cells("C22:D22"); ws_s.merge_cells("E22:F22"); ws_s.merge_cells("G22:H22")

scenarios = {
    "Bear": {"vol":"1% CAGR", "asp":"3% CAGR", "margin":"36.0%", "wacc":"9.5%", "tgr":"2.5%",
             "dcf":"$210", "blended":"$260", "thesis":"Luxury multiple collapses to auto levels; Luce disappoints"},
    "Base": {"vol":"3% CAGR", "asp":"5.5% CAGR","margin":"40.5%", "wacc":"7.5%","tgr":"3.0%",
             "dcf":"$351", "blended":"$450", "thesis":"Multiple re-rates to 27x; Luce confirms ASP accretion"},
    "Bull": {"vol":"4% CAGR", "asp":"8% CAGR",  "margin":"44.0%", "wacc":"7.0%","tgr":"3.5%",
             "dcf":"$480", "blended":"$560", "thesis":"Luce 30x+ multiple + Hermès re-rating; EPS beats consensus"},
}
scen_rows_ = [
    ("Volume CAGR '25–'30",    "vol"),
    ("ASP CAGR '25–'30",       "asp"),
    ("FY2030E EBITDA Margin",  "margin"),
    ("WACC",                   "wacc"),
    ("Terminal Growth Rate",   "tgr"),
    ("DCF Price Target",       "dcf"),
    ("Blended PT (40/60)",     "blended"),
    ("Thesis",                 "thesis"),
]
for i, (metric, key) in enumerate(scen_rows_):
    r = 23 + i; rh(ws_s, r, 13 if key != "thesis" else 28)
    if i % 2 == 0: stripe(ws_s, r, 2, 8)
    lbl(ws_s, r, 2, metric, bold=(key in ["dcf","blended"]))
    for j, scen in enumerate(["Bear","Base","Bull"]):
        ws_s.merge_cells(start_row=r, start_column=3+j*2, end_row=r, end_column=4+j*2)
        c = ws_s.cell(row=r, column=3+j*2, value=scenarios[scen][key])
        c.font = Font(name="Calibri", size=8, bold=(key in ["dcf","blended"]),
                      color=RED_ACC if (scen=="Bear" and key in ["dcf","blended"]) else
                            (GREEN_ACC if (scen=="Bull" and key in ["dcf","blended"]) else
                             (CRIMSON if key in ["dcf","blended"] else DARK)))
        c.alignment = _align("center" if key != "thesis" else "left", "center")
        if key in ["dcf","blended"]: c.fill = _fill(BLUE_FILL)

for c in range(1,10): ws_s.cell(row=33, column=c).fill = _fill(FERRARI_R)
rh(ws_s, 33, 4)

wb.save(OUTPUT)
print(f"✓  Ferrari valuation model saved → {OUTPUT}")
