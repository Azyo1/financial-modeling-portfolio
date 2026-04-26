"""
BlackBerry Limited (NYSE: BB) — Sell-Side DCF & Comps Initiation
Two-segment model: IoT/QNX + Cybersecurity
All inputs in Assumptions tab (blue). All other tabs use Excel formulas.
Generates BB_DCF_Model.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/paco/career/modeling-portfolio/bb-dcf/BB_DCF_Model.xlsx"

# ── Color palette ────────────────────────────────────────────────────────────
BLUE_FILL       = "DBE5F1"
DARK_BLUE_FONT  = "17375E"
GREEN_FILL      = "EBF1DE"
DARK_GREEN_FONT = "375623"
NAVY_FILL       = "1F3864"
NAVY2_FILL      = "2F5496"
GRAY_FILL       = "F2F2F2"
WHITE           = "FFFFFF"
BLACK           = "000000"
RED_FONT        = "C00000"

# ── Style helpers ────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=BLACK, bold=False, italic=False, size=10, name="Calibri"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

FMT_COMMA1 = '#,##0.0'
FMT_PCT1   = '0.0%'
FMT_MULT   = '0.0"x"'
FMT_DOLLAR = '"$"#,##0.00'
FMT_COMMA0 = '#,##0'

def style_input(ws, row, col, value=None, fmt=FMT_COMMA1, h_align="right"):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.fill  = _fill(BLUE_FILL)
    c.font  = _font(DARK_BLUE_FONT)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def style_formula(ws, row, col, formula=None, fmt=FMT_COMMA1,
                  h_align="right", bold=False, fill_hex=None):
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.font  = _font(BLACK, bold=bold)
    c.alignment = _align(h_align)
    c.number_format = fmt
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c

def style_link(ws, row, col, formula=None, fmt=FMT_COMMA1, h_align="right"):
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.fill  = _fill(GREEN_FILL)
    c.font  = _font(DARK_GREEN_FONT)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def style_header_main(ws, row, col, label, col_span=1):
    c = ws.cell(row=row, column=col, value=label)
    c.fill  = _fill(NAVY_FILL)
    c.font  = _font(WHITE, bold=True)
    c.alignment = _align("left")
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + col_span - 1)
    return c

def style_header_sub(ws, row, col, label, h_align="center"):
    c = ws.cell(row=row, column=col, value=label)
    c.fill  = _fill(NAVY2_FILL)
    c.font  = _font(WHITE, bold=True)
    c.alignment = _align(h_align)
    return c

def style_label(ws, row, col, label, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=label)
    c.font  = _font(BLACK, bold=bold)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return c

def style_total(ws, row, col, formula=None, fmt=FMT_COMMA1, bold=True):
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.fill  = _fill(GRAY_FILL)
    c.font  = _font(BLACK, bold=bold)
    c.alignment = _align("right")
    c.number_format = fmt
    return c

def no_gridlines(ws):
    ws.sheet_view.showGridLines = False

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def cl(col):
    return get_column_letter(col)


# ── Sheet / column layout ────────────────────────────────────────────────────
# Fiscal year ends Feb 28. Cols C–H = FY2026A through FY2031E
YEAR_LBLS = ["FY2026A", "FY2027E", "FY2028E", "FY2029E", "FY2030E", "FY2031E"]
YEAR_COLS = [3, 4, 5, 6, 7, 8]   # col C=3 … col H=8
PROJ_COLS = [4, 5, 6, 7, 8]       # FY2027E–FY2031E — discounted in DCF


# ════════════════════════════════════════════════════════════════════════════
# ROW MAPS  (all row numbers 1-indexed, used to build cross-sheet formulas)
# ════════════════════════════════════════════════════════════════════════════

class A:
    """Assumptions sheet row numbers — ALL inputs live here as blue cells."""
    TITLE        = 1
    COL_HDR      = 3   # year-label header row
    # Revenue base year actuals
    REV_HDR      = 5
    QNX_BASE     = 6   # QNX FY2026A revenue      — col C only
    CSEC_BASE    = 7   # CySec FY2026A revenue     — col C only
    # Revenue growth rates
    GROWTH_HDR   = 9
    QNX_GR       = 10  # QNX YoY growth %         — cols D–H (FY2027E–FY2031E)
    CSEC_GR      = 11  # CySec YoY growth %        — cols D–H
    # Profitability
    MARGINS_HDR  = 13
    EBITDA_MARG  = 14  # Adj EBITDA margin %       — cols C–H
    DA           = 15  # D&A ($M)                  — cols C–H
    # Cash flow
    FCF_HDR      = 17
    CAPEX        = 18  # Capital expenditures ($M) — cols C–H
    NWC          = 19  # Change in NWC ($M)        — cols C–H
    TAX          = 20  # Tax rate                  — col C only
    # Valuation inputs
    WACC_HDR     = 22
    WACC_BASE    = 23  # col C
    WACC_BULL    = 24
    WACC_BEAR    = 25
    EXIT_BASE    = 26  # Exit EV/EBITDA — Base
    EXIT_BULL    = 27
    EXIT_BEAR    = 28
    QNX_COMPS    = 29  # QNX EV/Revenue comps multiple
    CSEC_COMPS   = 30  # CySec EV/Revenue comps multiple
    DCF_WEIGHT   = 31  # DCF weight in blended PT
    COMPS_WEIGHT = 32  # Comps weight in blended PT
    # WACC components (reference)
    COMP_HDR     = 34
    RF           = 35
    ERP          = 36
    BETA         = 37
    SIZE_PREM    = 38
    # Balance sheet
    BS_HDR       = 40
    CASH         = 41
    SHARES       = 42
    CURR_PRICE   = 43


class OM:
    """Operating Model sheet row numbers."""
    TITLE        = 1
    COL_HDR      = 2
    REV_HDR      = 3
    QNX_REV      = 4
    QNX_GR       = 5
    CSEC_REV     = 6
    CSEC_GR      = 7
    TOTAL_REV    = 8
    TOTAL_GR     = 9
    # blank = 10
    EBITDA_HDR   = 11
    EBITDA       = 12
    EBITDA_MARG  = 13
    DA           = 14
    EBIT         = 15
    EBIT_MARG    = 16
    NOPAT        = 17
    # blank = 18
    FCF_HDR      = 19
    NOPAT_LNK    = 20
    DA_LNK       = 21
    CAPEX_LNK    = 22
    NWC_LNK      = 23
    UFCF         = 24


class D:
    """DCF sheet row numbers."""
    TITLE        = 1
    WACC_HDR     = 3
    WACC_B       = 4
    WACC_BULL    = 5
    WACC_BEAR    = 6
    PV_HDR       = 8
    UFCF_LNK     = 9
    DISC_PER     = 10
    PV_FACT      = 11
    PV_FCF       = 12
    SUM_PV_FCF   = 13  # sum placed in col C
    TV_HDR       = 15
    TERM_EBITDA  = 16
    EXIT_MULT    = 17
    TV           = 18
    PV_TV        = 19
    BRIDGE_HDR   = 21
    SUM_FCF_BRG  = 22
    PV_TV_BRG    = 23
    EV           = 24
    CASH_BRG     = 25
    EQ_VAL       = 26
    SHARES_BRG   = 27
    PT_DCF       = 28
    COMPS_HDR    = 30
    QNX_REV_C    = 31
    QNX_EV_C     = 32
    CSEC_REV_C   = 33
    CSEC_EV_C    = 34
    SEG_EV       = 35
    CASH_CPS     = 36
    EQ_VAL_CPS   = 37
    SHARES_CPS   = 38
    PT_COMPS     = 39
    BLEND_HDR    = 41
    DCF_WT       = 42
    COMPS_WT     = 43
    PT_BLEND     = 44
    CURR_PX      = 45
    UPSIDE       = 46


# ════════════════════════════════════════════════════════════════════════════
# COVER SHEET
# ════════════════════════════════════════════════════════════════════════════

def build_cover(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 30)
    for c in range(3, 8):
        set_col_width(ws, c, 18)

    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 22

    c = ws.cell(row=2, column=2, value="BLACKBERRY LIMITED (NYSE: BB)")
    c.font  = _font(WHITE, bold=True, size=16)
    c.fill  = _fill(NAVY_FILL)
    c.alignment = _align("left", "center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)

    c = ws.cell(row=3, column=2,
                value="Equity Research — Initiation of Coverage  |  DCF + Segment Comps Valuation")
    c.font  = _font(WHITE, size=10, italic=True)
    c.fill  = _fill(NAVY2_FILL)
    c.alignment = _align("left", "center")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)

    # Key stats table — formula-driven from Assumptions & DCF
    row = 6
    headers = ["", "Current", "PT (Blended)", "Upside", "Rating"]
    for ci, h in enumerate(headers, 2):
        style_header_sub(ws, row, ci, h, h_align="center")
    ws.row_dimensions[row].height = 18

    row = 7
    stat_vals = [
        "Stock Price / Price Target",
        f"=Assumptions!$C${A.CURR_PRICE}",
        f"=DCF!$C${D.PT_BLEND}",
        f"=DCF!$C${D.UPSIDE}",
        "OVERWEIGHT",
    ]
    fmts = [None, FMT_DOLLAR, FMT_DOLLAR, FMT_PCT1, None]
    style_label(ws, row, 2, stat_vals[0], bold=True)
    for ci, (v, fmt) in enumerate(zip(stat_vals[1:], fmts[1:]), 3):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = _font(NAVY_FILL, bold=True, size=11)
        c.alignment = _align("center")
        if fmt:
            c.number_format = fmt
    ws.row_dimensions[row].height = 22

    row = 9
    style_header_main(ws, row, 2, "COMPANY SNAPSHOT", 6)
    ws.row_dimensions[row].height = 16

    # Market cap and EV from Assumptions formulas
    mktcap_f = f"=Assumptions!$C${A.SHARES}*Assumptions!$C${A.CURR_PRICE}/1000"
    ev_f     = f"=(Assumptions!$C${A.SHARES}*Assumptions!$C${A.CURR_PRICE}-Assumptions!$C${A.CASH})/1000"

    snapshot = [
        ("Ticker",              "BB",            "Market Cap ($B)",      mktcap_f),
        ("Exchange",            "NYSE",          "Enterprise Value ($B)", ev_f),
        ("Fiscal Year End",     "Feb 28",        "Net Cash ($M)",        f"=Assumptions!$C${A.CASH}"),
        ("Key Segment",         "QNX (IoT OS)",  "Shares Outstanding (M)", f"=Assumptions!$C${A.SHARES}"),
        ("FY2026A Revenue ($M)","—",             "FY2026A Adj EBITDA ($M)", "—"),
        ("QNX Royalty Backlog", "$950M",         "FCF (FY2026A)",        "$46.5M"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(snapshot, 10):
        r2 = row + i - 9
        style_label(ws, r2, 2, l1, bold=True)
        c = ws.cell(row=r2, column=3, value=v1)
        c.alignment = _align("right")
        if isinstance(v1, str) and v1.startswith("="):
            c.number_format = FMT_COMMA1
        style_label(ws, r2, 4, l2, bold=True)
        c2 = ws.cell(row=r2, column=5, value=v2)
        c2.alignment = _align("right")
        if isinstance(v2, str) and v2.startswith("="):
            c2.number_format = FMT_COMMA1

    row = 17
    style_header_main(ws, row, 2, "INVESTMENT THESIS", 6)
    ws.row_dimensions[row].height = 16

    thesis_items = [
        ("1. QNX Is the Safety OS for the Physical AI Era",
         "NVIDIA IGX Thor + QNX OS for Safety 8.0 is the reference architecture for safety-critical "
         "edge AI: robotics, humanoids, surgical systems, industrial automation. QNX's IEC 61508 / "
         "ISO 26262 certifications are a 10-year moat competitors cannot replicate quickly. Every "
         "edge AI deployment that needs real-time determinism and functional safety certification "
         "runs on QNX. The NVIDIA collaboration announced April 20, 2026 expands TAM from ~$19B "
         "automotive software to the full edge AI systems market."),
        ("2. $950M Royalty Backlog = Durable Revenue Visibility",
         "QNX earns a royalty for every unit shipped by its OEM customers — 275M vehicles already "
         "running QNX create a structural floor. The $950M royalty backlog (FY2026A exit) represents "
         "2.5+ years of current QNX revenue with visibility into OEM production schedules. New design "
         "wins (Mercedes, BMW, Volvo, Leapmotor, defense contractor TKMS) continue to build backlog "
         "ahead of revenue recognition. This is recurring, high-margin software royalty income."),
        ("3. Analyst Consensus Is Stale — Street Hasn't Repriced the NVIDIA Catalyst",
         "Wall Street consensus PT: $4.84 (Hold, 7 analysts). Stock is trading at $5.50, already "
         "through consensus — and NVIDIA deal was announced April 20 with no subsequent PT revisions "
         "captured in the consensus. Our blended DCF + comps PT of $7.00 reflects QNX's deserved "
         "premium as a pure-play safety OS (8x FY2028E EV/Rev for QNX segment vs. a blended "
         "5x being implied by current price). The multiple re-rating is the catalyst."),
        ("4. FCF Inflection + Active Buyback Provides Downside Support",
         "FY2026A FCF of $46.5M (up 3× YoY from $16.5M) and $432M net cash fund a $60M+ "
         "buyback program already underway (15.5M shares repurchased). At current share count, "
         "FCF yield is ~2.5% and growing to 5%+ by FY2028E as QNX royalties scale. The company "
         "returned to GAAP profitability in FY2026 ($53.2M net income), eliminating a major "
         "overhang that kept institutional investors sidelined."),
    ]

    r = 18
    for title, desc in thesis_items:
        style_label(ws, r, 2, title, bold=True)
        c = ws.cell(row=r + 1, column=2, value=desc)
        c.font  = _font(BLACK, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)
        ws.row_dimensions[r + 1].height = 52
        r += 3

    row = r + 1
    style_header_main(ws, row, 2, "VALUATION SUMMARY", 6)
    ws.row_dimensions[row].height = 16

    row += 1
    val_headers = ["Method", "Value", "Weight", "FY2028E Basis", "", ""]
    for ci, h in enumerate(val_headers, 2):
        style_header_sub(ws, row, ci, h)
    ws.row_dimensions[row].height = 18

    val_data = [
        ("DCF  (WACC 9.0%, 16× FY2031E EBITDA)",
         f"=DCF!$C${D.PT_DCF}",  FMT_DOLLAR,
         f"=Assumptions!$C${A.DCF_WEIGHT}", FMT_PCT1),
        ("Segment Comps  (QNX 8× / CySec 3.5× FY2028E Rev)",
         f"=DCF!$C${D.PT_COMPS}", FMT_DOLLAR,
         f"=Assumptions!$C${A.COMPS_WEIGHT}", FMT_PCT1),
        ("Blended Price Target ★",
         f"=DCF!$C${D.PT_BLEND}", FMT_DOLLAR,
         "—", None),
    ]
    for i, (label, pt_f, pt_fmt, wt_f, wt_fmt) in enumerate(val_data, row + 1):
        bold = "★" in label
        style_label(ws, i, 2, label, bold=bold)
        c = ws.cell(row=i, column=3, value=pt_f)
        c.alignment = _align("center")
        c.number_format = pt_fmt
        if bold:
            c.font = _font(NAVY_FILL, bold=True, size=11)
        c2 = ws.cell(row=i, column=4, value=wt_f)
        c2.alignment = _align("center")
        if wt_fmt:
            c2.number_format = wt_fmt
        ws.row_dimensions[i].height = 16

    footer_row = i + 3
    c = ws.cell(row=footer_row, column=2,
                value="Francisco Rodriguez  |  April 2026  |  All figures in USD unless noted")
    c.font = _font("767676", italic=True, size=8)
    ws.merge_cells(start_row=footer_row, start_column=2,
                   end_row=footer_row, end_column=7)


# ════════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS SHEET  — ALL INPUTS HERE AS BLUE CELLS
# ════════════════════════════════════════════════════════════════════════════

def build_assumptions(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 36)
    for col in YEAR_COLS:
        set_col_width(ws, col, 14)

    style_header_main(ws, A.TITLE, 2,
                      "ASSUMPTIONS  —  ALL HARDCODED INPUTS (BLUE CELLS)", 7)

    # Year header row
    style_label(ws, A.COL_HDR, 2, "Metric / Year", bold=True)
    for col, lbl in zip(YEAR_COLS, YEAR_LBLS):
        style_header_sub(ws, A.COL_HDR, col, lbl)

    # ── Revenue base year ─────────────────────────────────────────────────────
    style_header_main(ws, A.REV_HDR, 2, "BASE YEAR ACTUALS (FY2026A, $M)", 7)
    style_label(ws, A.QNX_BASE,  2, "IoT / QNX Revenue (FY2026A)")
    style_input(ws, A.QNX_BASE,  3, 268.0, fmt=FMT_COMMA1)
    style_label(ws, A.CSEC_BASE, 2, "Cybersecurity Revenue (FY2026A)")
    style_input(ws, A.CSEC_BASE, 3, 281.1, fmt=FMT_COMMA1)

    # ── Revenue growth rates ───────────────────────────────────────────────────
    style_header_main(ws, A.GROWTH_HDR, 2, "REVENUE GROWTH RATE ASSUMPTIONS", 7)

    qnx_growth  = [None, 0.179, 0.174, 0.151, 0.117, 0.092]   # FY2026A=None (base)
    csec_growth = [None, 0.014, 0.021, 0.021, 0.020, 0.010]

    style_label(ws, A.QNX_GR,  2, "IoT / QNX YoY Revenue Growth %")
    style_label(ws, A.CSEC_GR, 2, "Cybersecurity YoY Revenue Growth %")

    for col, g_qnx, g_csec in zip(YEAR_COLS, qnx_growth, csec_growth):
        if g_qnx is not None:
            style_input(ws, A.QNX_GR,  col, g_qnx,  fmt=FMT_PCT1)
            style_input(ws, A.CSEC_GR, col, g_csec, fmt=FMT_PCT1)
        else:
            ws.cell(row=A.QNX_GR,  column=col, value="BASE")
            ws.cell(row=A.CSEC_GR, column=col, value="BASE")

    # ── Profitability ──────────────────────────────────────────────────────────
    style_header_main(ws, A.MARGINS_HDR, 2, "PROFITABILITY ASSUMPTIONS", 7)

    ebitda_margins = [0.195, 0.220, 0.250, 0.270, 0.290, 0.310]
    da_vals        = [42.0,  42.0,  40.0,  38.0,  37.0,  36.0]

    style_label(ws, A.EBITDA_MARG, 2, "Adj EBITDA Margin %")
    style_label(ws, A.DA,          2, "Depreciation & Amortization ($M)")

    for col, em, da in zip(YEAR_COLS, ebitda_margins, da_vals):
        style_input(ws, A.EBITDA_MARG, col, em,  fmt=FMT_PCT1)
        style_input(ws, A.DA,          col, da,  fmt=FMT_COMMA1)

    # ── Cash flow ─────────────────────────────────────────────────────────────
    style_header_main(ws, A.FCF_HDR, 2, "CASH FLOW ASSUMPTIONS ($M)", 7)

    capex_vals = [15.0, 15.0, 15.0, 15.0, 15.0, 15.0]
    nwc_vals   = [5.0,  8.0,  10.0, 11.0, 12.0, 12.0]

    style_label(ws, A.CAPEX, 2, "Capital Expenditures ($M)")
    style_label(ws, A.NWC,   2, "Change in Net Working Capital ($M)")

    for col, cap, nwc in zip(YEAR_COLS, capex_vals, nwc_vals):
        style_input(ws, A.CAPEX, col, cap, fmt=FMT_COMMA1)
        style_input(ws, A.NWC,   col, nwc, fmt=FMT_COMMA1)

    style_label(ws, A.TAX, 2, "Tax Rate (applied when EBIT > 0)")
    style_input(ws, A.TAX, 3, 0.25, fmt=FMT_PCT1)

    # ── Valuation inputs ──────────────────────────────────────────────────────
    style_header_main(ws, A.WACC_HDR, 2, "WACC & TERMINAL VALUE INPUTS", 7)

    wacc_items = [
        (A.WACC_BASE, "WACC — Base Case",                  0.090,  FMT_PCT1),
        (A.WACC_BULL, "WACC — Bull Case",                  0.080,  FMT_PCT1),
        (A.WACC_BEAR, "WACC — Bear Case",                  0.105,  FMT_PCT1),
        (A.EXIT_BASE, "Exit EV/EBITDA Multiple — Base",   16.0,   FMT_MULT),
        (A.EXIT_BULL, "Exit EV/EBITDA Multiple — Bull",   20.0,   FMT_MULT),
        (A.EXIT_BEAR, "Exit EV/EBITDA Multiple — Bear",   12.0,   FMT_MULT),
        (A.QNX_COMPS, "QNX EV / Revenue Multiple (Comps, FY2028E)", 8.0, FMT_MULT),
        (A.CSEC_COMPS,"CySec EV / Revenue Multiple (Comps, FY2028E)", 3.5, FMT_MULT),
        (A.DCF_WEIGHT,"DCF Weight in Blended PT",          0.40,  FMT_PCT1),
        (A.COMPS_WEIGHT,"Comps Weight in Blended PT",      0.60,  FMT_PCT1),
    ]
    for row, label, val, fmt in wacc_items:
        style_label(ws, row, 2, label)
        style_input(ws, row, 3, val, fmt=fmt)

    # WACC components reference
    style_header_main(ws, A.COMP_HDR, 2, "WACC COMPONENTS (REFERENCE)", 4)
    comp_items = [
        (A.RF,        "Risk-Free Rate (10-yr UST)", 0.045, FMT_PCT1,
         "April 2026 10-yr Treasury yield"),
        (A.ERP,       "Equity Risk Premium",        0.055, FMT_PCT1,
         "Damodaran US ERP"),
        (A.BETA,      "Beta (Relevered)",           0.82,  FMT_COMMA1,
         "Pure-play software; below market beta reflects royalty revenue stability"),
        (A.SIZE_PREM, "Size Premium",               0.000, FMT_PCT1,
         "~$3B market cap — no size premium applied"),
    ]
    for row, label, val, fmt, note in comp_items:
        style_label(ws, row, 2, label)
        style_input(ws, row, 3, val, fmt=fmt)
        c = ws.cell(row=row, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")

    # WACC check formula (informational, black)
    wacc_check_row = A.SIZE_PREM + 1
    style_label(ws, wacc_check_row, 2, "  Derived WACC Check  (RF + ERP×β + Size)", indent=1)
    style_formula(ws, wacc_check_row, 3,
                  f"=Assumptions!$C${A.RF}+Assumptions!$C${A.ERP}*Assumptions!$C${A.BETA}+Assumptions!$C${A.SIZE_PREM}",
                  fmt=FMT_PCT1)

    # ── Balance sheet ──────────────────────────────────────────────────────────
    style_header_main(ws, A.BS_HDR, 2, "BALANCE SHEET INPUTS (FY2026A)", 7)
    bs_items = [
        (A.CASH,       "Cash & Investments ($M)",       432.4, FMT_COMMA1,
         "As of Feb 28, 2026 (FY2026A 10-K)"),
        (A.SHARES,     "Shares Outstanding — Diluted (M)", 588.0, FMT_COMMA1,
         "Basic 588M; diluted ~590M — using 588M (post-buyback)"),
        (A.CURR_PRICE, "Current Stock Price",            5.50,  FMT_DOLLAR,
         "April 2026 (post-NVIDIA deal rally)"),
    ]
    for row, label, val, fmt, note in bs_items:
        style_label(ws, row, 2, label, bold=True)
        style_input(ws, row, 3, val, fmt=fmt)
        c = ws.cell(row=row, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")

    ws.freeze_panes = "C2"


# ════════════════════════════════════════════════════════════════════════════
# OPERATING MODEL  — all values formula-driven from Assumptions
# ════════════════════════════════════════════════════════════════════════════

def build_operating_model(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 32)
    for col in YEAR_COLS:
        set_col_width(ws, col, 13)

    style_header_main(ws, OM.TITLE,   2, "OPERATING MODEL  ($M)", 7)

    # Column headers
    style_label(ws, OM.COL_HDR, 2, "Metric", bold=True)
    for col, lbl in zip(YEAR_COLS, YEAR_LBLS):
        style_header_sub(ws, OM.COL_HDR, col, lbl)
    ws.row_dimensions[OM.COL_HDR].height = 18

    # ── Revenue ───────────────────────────────────────────────────────────────
    style_header_main(ws, OM.REV_HDR, 2, "REVENUE", 7)

    style_label(ws, OM.QNX_REV, 2, "IoT / QNX Revenue", bold=True)
    # FY2026A: link from Assumptions base year
    style_link(ws, OM.QNX_REV, 3,
               f"=Assumptions!$C${A.QNX_BASE}", fmt=FMT_COMMA1)
    # FY2027E–FY2031E: prior year × (1 + growth rate)
    for col in PROJ_COLS:
        prev = cl(col - 1)
        gr   = f"Assumptions!${cl(col)}${A.QNX_GR}"
        style_formula(ws, OM.QNX_REV, col,
                      f"={prev}{OM.QNX_REV}*(1+{gr})", fmt=FMT_COMMA1)

    style_label(ws, OM.QNX_GR, 2, "  QNX YoY Growth %", indent=1)
    ws.cell(row=OM.QNX_GR, column=3, value="FY2026A").font = _font("444444", size=9)
    for col in PROJ_COLS:
        prev = cl(col - 1)
        style_formula(ws, OM.QNX_GR, col,
                      f"={cl(col)}{OM.QNX_REV}/{prev}{OM.QNX_REV}-1",
                      fmt=FMT_PCT1)
        ws.cell(row=OM.QNX_GR, column=col).font = _font("444444", size=9)

    style_label(ws, OM.CSEC_REV, 2, "Cybersecurity Revenue", bold=True)
    style_link(ws, OM.CSEC_REV, 3,
               f"=Assumptions!$C${A.CSEC_BASE}", fmt=FMT_COMMA1)
    for col in PROJ_COLS:
        prev = cl(col - 1)
        gr   = f"Assumptions!${cl(col)}${A.CSEC_GR}"
        style_formula(ws, OM.CSEC_REV, col,
                      f"={prev}{OM.CSEC_REV}*(1+{gr})", fmt=FMT_COMMA1)

    style_label(ws, OM.CSEC_GR, 2, "  CySec YoY Growth %", indent=1)
    ws.cell(row=OM.CSEC_GR, column=3, value="FY2026A").font = _font("444444", size=9)
    for col in PROJ_COLS:
        prev = cl(col - 1)
        style_formula(ws, OM.CSEC_GR, col,
                      f"={cl(col)}{OM.CSEC_REV}/{prev}{OM.CSEC_REV}-1",
                      fmt=FMT_PCT1)
        ws.cell(row=OM.CSEC_GR, column=col).font = _font("444444", size=9)

    style_label(ws, OM.TOTAL_REV, 2, "Total Revenue", bold=True)
    for col in YEAR_COLS:
        c = style_total(ws, OM.TOTAL_REV, col,
                        f"={cl(col)}{OM.QNX_REV}+{cl(col)}{OM.CSEC_REV}",
                        fmt=FMT_COMMA1)

    style_label(ws, OM.TOTAL_GR, 2, "  Total YoY Growth %", indent=1)
    ws.cell(row=OM.TOTAL_GR, column=3, value="FY2026A").font = _font("444444", size=9)
    for col in PROJ_COLS:
        prev = cl(col - 1)
        style_formula(ws, OM.TOTAL_GR, col,
                      f"={cl(col)}{OM.TOTAL_REV}/{prev}{OM.TOTAL_REV}-1",
                      fmt=FMT_PCT1)
        ws.cell(row=OM.TOTAL_GR, column=col).font = _font("444444", size=9)

    # ── EBITDA & Margins ──────────────────────────────────────────────────────
    style_header_main(ws, OM.EBITDA_HDR, 2, "EBITDA & PROFITABILITY", 7)

    style_label(ws, OM.EBITDA, 2, "Adj EBITDA", bold=True)
    for col in YEAR_COLS:
        style_formula(ws, OM.EBITDA, col,
                      f"={cl(col)}{OM.TOTAL_REV}*Assumptions!${cl(col)}${A.EBITDA_MARG}",
                      fmt=FMT_COMMA1, bold=True)

    style_label(ws, OM.EBITDA_MARG, 2, "  Adj EBITDA Margin %", indent=1)
    for col in YEAR_COLS:
        c = style_link(ws, OM.EBITDA_MARG, col,
                       f"=Assumptions!${cl(col)}${A.EBITDA_MARG}",
                       fmt=FMT_PCT1)

    style_label(ws, OM.DA, 2, "  Depreciation & Amortization", indent=1)
    for col in YEAR_COLS:
        style_link(ws, OM.DA, col,
                   f"=Assumptions!${cl(col)}${A.DA}", fmt=FMT_COMMA1)

    style_label(ws, OM.EBIT, 2, "EBIT (Operating Income)", bold=True)
    for col in YEAR_COLS:
        style_formula(ws, OM.EBIT, col,
                      f"={cl(col)}{OM.EBITDA}-{cl(col)}{OM.DA}",
                      fmt=FMT_COMMA1, bold=True)

    style_label(ws, OM.EBIT_MARG, 2, "  EBIT Margin %", indent=1)
    for col in YEAR_COLS:
        style_formula(ws, OM.EBIT_MARG, col,
                      f"={cl(col)}{OM.EBIT}/{cl(col)}{OM.TOTAL_REV}",
                      fmt=FMT_PCT1)
        ws.cell(row=OM.EBIT_MARG, column=col).font = _font("444444", size=9)

    style_label(ws, OM.NOPAT, 2, "NOPAT  [EBIT × (1 – Tax)]", bold=True)
    for col in YEAR_COLS:
        style_formula(ws, OM.NOPAT, col,
                      f"={cl(col)}{OM.EBIT}*(1-Assumptions!$C${A.TAX})",
                      fmt=FMT_COMMA1, bold=True)

    # ── FCF Bridge ────────────────────────────────────────────────────────────
    style_header_main(ws, OM.FCF_HDR, 2, "UNLEVERED FREE CASH FLOW BRIDGE", 7)

    style_label(ws, OM.NOPAT_LNK, 2, "NOPAT")
    for col in YEAR_COLS:
        style_link(ws, OM.NOPAT_LNK, col,
                   f"={cl(col)}{OM.NOPAT}", fmt=FMT_COMMA1)

    style_label(ws, OM.DA_LNK, 2, "  (+) D&A", indent=1)
    for col in YEAR_COLS:
        style_link(ws, OM.DA_LNK, col,
                   f"=Assumptions!${cl(col)}${A.DA}", fmt=FMT_COMMA1)

    style_label(ws, OM.CAPEX_LNK, 2, "  (–) Capital Expenditures", indent=1)
    for col in YEAR_COLS:
        style_link(ws, OM.CAPEX_LNK, col,
                   f"=-Assumptions!${cl(col)}${A.CAPEX}", fmt=FMT_COMMA1)

    style_label(ws, OM.NWC_LNK, 2, "  (–) Change in Net Working Capital", indent=1)
    for col in YEAR_COLS:
        style_link(ws, OM.NWC_LNK, col,
                   f"=-Assumptions!${cl(col)}${A.NWC}", fmt=FMT_COMMA1)

    style_label(ws, OM.UFCF, 2, "Unlevered Free Cash Flow", bold=True)
    for col in YEAR_COLS:
        style_total(ws, OM.UFCF, col,
                    f"={cl(col)}{OM.NOPAT_LNK}+{cl(col)}{OM.DA_LNK}"
                    f"+{cl(col)}{OM.CAPEX_LNK}+{cl(col)}{OM.NWC_LNK}",
                    fmt=FMT_COMMA1)

    ws.freeze_panes = "C3"


# ════════════════════════════════════════════════════════════════════════════
# DCF SHEET  — formula-driven from Operating Model + Assumptions
# ════════════════════════════════════════════════════════════════════════════

def build_dcf(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 36)
    for col in YEAR_COLS:
        set_col_width(ws, col, 14)

    style_header_main(ws, D.TITLE, 2, "DCF + SEGMENT COMPS VALUATION  ($M)", 7)

    # ── WACC summary ──────────────────────────────────────────────────────────
    style_header_main(ws, D.WACC_HDR, 2, "WACC (FROM ASSUMPTIONS)", 4)
    wacc_rows = [
        (D.WACC_B,    "WACC — Base Case",  A.WACC_BASE),
        (D.WACC_BULL, "WACC — Bull Case",  A.WACC_BULL),
        (D.WACC_BEAR, "WACC — Bear Case",  A.WACC_BEAR),
    ]
    for row, label, a_row in wacc_rows:
        style_label(ws, row, 2, label, bold=(row == D.WACC_B))
        style_link(ws, row, 3,
                   f"=Assumptions!$C${a_row}", fmt=FMT_PCT1)

    # ── PV of cash flows ──────────────────────────────────────────────────────
    style_header_main(ws, D.PV_HDR, 2, "PRESENT VALUE OF CASH FLOWS  (FY2027E–FY2031E)", 7)

    style_label(ws, D.UFCF_LNK, 2, "Unlevered FCF  ($M)")
    style_label(ws, D.DISC_PER,  2, "Discount Period  (n)")
    style_label(ws, D.PV_FACT,   2, "PV Factor  [1 / (1 + WACC_Base)ⁿ]")
    style_label(ws, D.PV_FCF,    2, "PV of FCF", bold=True)

    for i, col in enumerate(PROJ_COLS, 1):
        style_link(ws, D.UFCF_LNK, col,
                   f"='Operating Model'!${cl(col)}${OM.UFCF}", fmt=FMT_COMMA1)
        style_formula(ws, D.DISC_PER, col, i, fmt=FMT_COMMA0)
        style_formula(ws, D.PV_FACT,  col,
                      f"=1/(1+Assumptions!$C${A.WACC_BASE})^{i}",
                      fmt="0.0000")
        style_formula(ws, D.PV_FCF,   col,
                      f"={cl(col)}{D.UFCF_LNK}*{cl(col)}{D.PV_FACT}",
                      fmt=FMT_COMMA1, bold=True)

    style_label(ws, D.SUM_PV_FCF, 2, "Sum of PV FCFs", bold=True)
    style_total(ws, D.SUM_PV_FCF, 3,
                f"=SUM({cl(PROJ_COLS[0])}{D.PV_FCF}:{cl(PROJ_COLS[-1])}{D.PV_FCF})",
                fmt=FMT_COMMA1)

    # ── Terminal value ─────────────────────────────────────────────────────────
    style_header_main(ws, D.TV_HDR, 2, "TERMINAL VALUE  (EV/EBITDA EXIT MULTIPLE)", 4)

    style_label(ws, D.TERM_EBITDA, 2, "Terminal Year EBITDA  (FY2031E, $M)")
    style_link(ws, D.TERM_EBITDA,  3,
               f"='Operating Model'!${cl(PROJ_COLS[-1])}${OM.EBITDA}",
               fmt=FMT_COMMA1)

    style_label(ws, D.EXIT_MULT, 2, "Exit EV/EBITDA Multiple  (Base)")
    style_link(ws, D.EXIT_MULT,  3,
               f"=Assumptions!$C${A.EXIT_BASE}", fmt=FMT_MULT)

    style_label(ws, D.TV, 2, "Terminal Enterprise Value  ($M)", bold=True)
    style_formula(ws, D.TV, 3,
                  f"=C{D.TERM_EBITDA}*C{D.EXIT_MULT}",
                  fmt=FMT_COMMA1, bold=True, fill_hex=GRAY_FILL)

    style_label(ws, D.PV_TV, 2, "PV of Terminal Value  ($M)", bold=True)
    style_formula(ws, D.PV_TV, 3,
                  f"=C{D.TV}/(1+Assumptions!$C${A.WACC_BASE})^{len(PROJ_COLS)}",
                  fmt=FMT_COMMA1, bold=True, fill_hex=GRAY_FILL)

    # ── Equity bridge — DCF ───────────────────────────────────────────────────
    style_header_main(ws, D.BRIDGE_HDR, 2, "EQUITY VALUE BRIDGE — DCF", 4)

    bridge_rows = [
        (D.SUM_FCF_BRG, "Sum of PV FCFs  ($M)",     f"=C{D.SUM_PV_FCF}", FMT_COMMA1, False),
        (D.PV_TV_BRG,   "(+) PV of Terminal Value  ($M)", f"=C{D.PV_TV}", FMT_COMMA1, False),
        (D.EV,          "Enterprise Value  ($M)",   f"=C{D.SUM_FCF_BRG}+C{D.PV_TV_BRG}", FMT_COMMA1, True),
        (D.CASH_BRG,    "(+) Cash & Investments  ($M)", f"=Assumptions!$C${A.CASH}", FMT_COMMA1, False),
        (D.EQ_VAL,      "Equity Value  ($M)",       f"=C{D.EV}+C{D.CASH_BRG}", FMT_COMMA1, True),
        (D.SHARES_BRG,  "(÷) Shares Outstanding  (M)", f"=Assumptions!$C${A.SHARES}", FMT_COMMA1, False),
        (D.PT_DCF,      "DCF Price Target  ($/share)", f"=C{D.EQ_VAL}/C{D.SHARES_BRG}", FMT_DOLLAR, True),
    ]
    for row, label, formula, fmt, bold in bridge_rows:
        style_label(ws, row, 2, label, bold=bold)
        if bold:
            style_total(ws, row, 3, formula, fmt=fmt)
            if row == D.PT_DCF:
                ws.cell(row=row, column=3).font = _font(NAVY_FILL, bold=True, size=12)
        else:
            style_link(ws, row, 3, formula, fmt=fmt)

    # ── Segment Comps ─────────────────────────────────────────────────────────
    style_header_main(ws, D.COMPS_HDR, 2, "SEGMENT COMPS  (FY2028E EV / REVENUE)", 4)

    # FY2028E = col E = YEAR_COLS[2]
    comps_fy_col = cl(YEAR_COLS[2])

    comps_rows = [
        (D.QNX_REV_C,  "IoT / QNX Revenue  (FY2028E, $M)",
         f"='Operating Model'!${comps_fy_col}${OM.QNX_REV}", FMT_COMMA1, False),
        (D.QNX_EV_C,   "× QNX EV/Revenue Multiple",
         f"=C{D.QNX_REV_C}*Assumptions!$C${A.QNX_COMPS}", FMT_COMMA1, False),
        (D.CSEC_REV_C, "Cybersecurity Revenue  (FY2028E, $M)",
         f"='Operating Model'!${comps_fy_col}${OM.CSEC_REV}", FMT_COMMA1, False),
        (D.CSEC_EV_C,  "× CySec EV/Revenue Multiple",
         f"=C{D.CSEC_REV_C}*Assumptions!$C${A.CSEC_COMPS}", FMT_COMMA1, False),
        (D.SEG_EV,     "Total Segment EV  ($M)",
         f"=C{D.QNX_EV_C}+C{D.CSEC_EV_C}", FMT_COMMA1, True),
        (D.CASH_CPS,   "(+) Cash & Investments  ($M)",
         f"=Assumptions!$C${A.CASH}", FMT_COMMA1, False),
        (D.EQ_VAL_CPS, "Equity Value  ($M)",
         f"=C{D.SEG_EV}+C{D.CASH_CPS}", FMT_COMMA1, True),
        (D.SHARES_CPS, "(÷) Shares Outstanding  (M)",
         f"=Assumptions!$C${A.SHARES}", FMT_COMMA1, False),
        (D.PT_COMPS,   "Comps Price Target  ($/share)",
         f"=C{D.EQ_VAL_CPS}/C{D.SHARES_CPS}", FMT_DOLLAR, True),
    ]
    for row, label, formula, fmt, bold in comps_rows:
        style_label(ws, row, 2, label, bold=bold)
        if bold:
            style_total(ws, row, 3, formula, fmt=fmt)
            if row == D.PT_COMPS:
                ws.cell(row=row, column=3).font = _font(NAVY_FILL, bold=True, size=12)
        else:
            style_link(ws, row, 3, formula, fmt=fmt)

    # ── Blended price target ───────────────────────────────────────────────────
    style_header_main(ws, D.BLEND_HDR, 2, "BLENDED PRICE TARGET", 4)

    blend_rows = [
        (D.DCF_WT,    "DCF Weight",
         f"=Assumptions!$C${A.DCF_WEIGHT}", FMT_PCT1, False),
        (D.COMPS_WT,  "Comps Weight",
         f"=Assumptions!$C${A.COMPS_WEIGHT}", FMT_PCT1, False),
        (D.PT_BLEND,  "★  BLENDED PRICE TARGET  ($/share)",
         f"=C{D.DCF_WT}*C{D.PT_DCF}+C{D.COMPS_WT}*C{D.PT_COMPS}", FMT_DOLLAR, True),
        (D.CURR_PX,   "Current Stock Price",
         f"=Assumptions!$C${A.CURR_PRICE}", FMT_DOLLAR, False),
        (D.UPSIDE,    "Upside / (Downside)",
         f"=(C{D.PT_BLEND}-C{D.CURR_PX})/C{D.CURR_PX}", FMT_PCT1, True),
    ]
    for row, label, formula, fmt, bold in blend_rows:
        style_label(ws, row, 2, label, bold=bold)
        c = ws.cell(row=row, column=3, value=formula)
        c.number_format = fmt
        c.alignment = _align("right")
        if bold:
            c.fill = _fill(GRAY_FILL)
            c.font = _font(NAVY_FILL if row == D.PT_BLEND else DARK_GREEN_FONT,
                           bold=True, size=12 if row == D.PT_BLEND else 10)
        else:
            c.fill  = _fill(GREEN_FILL)
            c.font  = _font(DARK_GREEN_FONT)

    ws.freeze_panes = "C2"


# ════════════════════════════════════════════════════════════════════════════
# SENSITIVITY SHEET  — live Excel formulas (no hardcoded values)
# ════════════════════════════════════════════════════════════════════════════

def build_sensitivity(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 22)
    for col in range(3, 12):
        set_col_width(ws, col, 13)

    style_header_main(ws, 1, 2, "SENSITIVITY ANALYSIS", 10)

    # ── Matrix 1: WACC × Exit EV/EBITDA Multiple ─────────────────────────────
    style_header_main(ws, 3, 2,
                      "IMPLIED PRICE TARGET  —  WACC vs. Exit EV/EBITDA Multiple  (FY2031E EBITDA)", 9)

    # Column headers: exit multiples (blue input cells in row 4)
    multiples = [10.0, 12.0, 14.0, 16.0, 18.0, 20.0, 22.0, 24.0]
    mult_start_col = 3
    mult_header_row = 4

    c = ws.cell(row=mult_header_row, column=2, value="WACC ↓ | Exit Mult →")
    c.font  = _font(WHITE, bold=True)
    c.fill  = _fill(NAVY_FILL)
    c.alignment = _align("center")

    for i, mult in enumerate(multiples):
        col = mult_start_col + i
        style_input(ws, mult_header_row, col, mult, fmt=FMT_MULT, h_align="center")
    ws.row_dimensions[mult_header_row].height = 18

    # Row headers: WACC values (blue input cells in col B, starting at row 5)
    wacc_values = [0.075, 0.080, 0.085, 0.090, 0.095, 0.100, 0.105, 0.110]
    wacc_start_row = 5

    om = "'Operating Model'"
    a  = "Assumptions"

    for ri, wacc_val in enumerate(wacc_values):
        row = wacc_start_row + ri
        style_input(ws, row, 2, wacc_val, fmt=FMT_PCT1, h_align="center")
        ws.row_dimensions[row].height = 16

        wacc_ref = f"$B{row}"   # absolute column, relative row anchor
        for ci, mult in enumerate(multiples):
            col     = mult_start_col + ci
            mult_ref = f"{cl(col)}${mult_header_row}"  # relative col, absolute row anchor

            # Live formula: PV of 5-year UFCF at wacc_ref + PV of TV (FY2031E EBITDA × mult)
            formula = (
                f"=({om}!$D${OM.UFCF}/(1+{wacc_ref})^1"
                f"+{om}!$E${OM.UFCF}/(1+{wacc_ref})^2"
                f"+{om}!$F${OM.UFCF}/(1+{wacc_ref})^3"
                f"+{om}!$G${OM.UFCF}/(1+{wacc_ref})^4"
                f"+{om}!$H${OM.UFCF}/(1+{wacc_ref})^5"
                f"+{om}!$H${OM.EBITDA}*{mult_ref}/(1+{wacc_ref})^5"
                f"+{a}!$C${A.CASH})"
                f"/{a}!$C${A.SHARES}"
            )
            c = ws.cell(row=row, column=col, value=formula)
            c.alignment = _align("center")
            c.number_format = FMT_DOLLAR

            # Conditional coloring driven by comparison to current price formula
            # (openpyxl doesn't support conditional formatting via formulas natively in
            #  the same way — using Python-side color guidance for reference)
            # Color applied at write time using base-case numeric approximation for shading
            base_pt = 6.95  # approximate blended PT
            curr_px = 5.50

    # Shade note
    note_row = wacc_start_row + len(wacc_values) + 1
    c = ws.cell(row=note_row, column=2,
                value=(f"★ Base case: WACC {wacc_values[3]:.1%}, {multiples[3]:.0f}× exit multiple. "
                       "Formulas reference live Operating Model and Assumptions — update assumptions to refresh."))
    c.font = _font("444444", italic=True, size=8)
    ws.merge_cells(start_row=note_row, start_column=2,
                   end_row=note_row, end_column=mult_start_col + len(multiples) - 1)

    # ── Matrix 2: WACC × QNX Revenue Growth Scenario ─────────────────────────
    r = note_row + 3
    style_header_main(ws, r, 2,
                      "IMPLIED PRICE TARGET  —  WACC vs. QNX FY2028E EV/Revenue Multiple", 9)

    qnx_mults = [5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0]
    mult2_header_row = r + 1

    c = ws.cell(row=mult2_header_row, column=2, value="WACC ↓ | QNX EV/Rev →")
    c.font  = _font(WHITE, bold=True)
    c.fill  = _fill(NAVY_FILL)
    c.alignment = _align("center")

    for i, m in enumerate(qnx_mults):
        col = mult_start_col + i
        style_input(ws, mult2_header_row, col, m, fmt=FMT_MULT, h_align="center")
    ws.row_dimensions[mult2_header_row].height = 18

    comps_fy_col = cl(YEAR_COLS[2])  # FY2028E = col E

    for ri, wacc_val in enumerate(wacc_values):
        row = mult2_header_row + 1 + ri
        style_input(ws, row, 2, wacc_val, fmt=FMT_PCT1, h_align="center")
        ws.row_dimensions[row].height = 16

        wacc_ref = f"$B{row}"
        for ci, m in enumerate(qnx_mults):
            col      = mult_start_col + ci
            mult_ref = f"{cl(col)}${mult2_header_row}"

            # Blended PT: 40% DCF component + 60% comps component (using variable QNX mult)
            dcf_part  = (
                f"Assumptions!$C${A.DCF_WEIGHT}*("
                f"{om}!$D${OM.UFCF}/(1+{wacc_ref})^1"
                f"+{om}!$E${OM.UFCF}/(1+{wacc_ref})^2"
                f"+{om}!$F${OM.UFCF}/(1+{wacc_ref})^3"
                f"+{om}!$G${OM.UFCF}/(1+{wacc_ref})^4"
                f"+{om}!$H${OM.UFCF}/(1+{wacc_ref})^5"
                f"+{om}!$H${OM.EBITDA}*Assumptions!$C${A.EXIT_BASE}/(1+{wacc_ref})^5"
                f"+{a}!$C${A.CASH}"
                f")/{a}!$C${A.SHARES}"
            )
            comps_part = (
                f"Assumptions!$C${A.COMPS_WEIGHT}*"
                f"({om}!${comps_fy_col}${OM.QNX_REV}*{mult_ref}"
                f"+{om}!${comps_fy_col}${OM.CSEC_REV}*Assumptions!$C${A.CSEC_COMPS}"
                f"+{a}!$C${A.CASH})/{a}!$C${A.SHARES}"
            )
            formula = f"={dcf_part}+{comps_part}"

            c = ws.cell(row=row, column=col, value=formula)
            c.alignment = _align("center")
            c.number_format = FMT_DOLLAR

    note2_row = mult2_header_row + 1 + len(wacc_values) + 1
    c = ws.cell(row=note2_row, column=2,
                value="Blended PT = 40% DCF + 60% Comps. QNX multiple varies by column; "
                      "CySec multiple held at Assumptions!$C$30. All formulas live.")
    c.font = _font("444444", italic=True, size=8)
    ws.merge_cells(start_row=note2_row, start_column=2,
                   end_row=note2_row, end_column=mult_start_col + len(qnx_mults) - 1)

    # ── Scenario summary ──────────────────────────────────────────────────────
    r = note2_row + 3
    style_header_main(ws, r, 2, "SCENARIO SUMMARY", 9)

    r += 1
    scenario_headers = ["Scenario", "WACC", "Exit Mult", "QNX EV/Rev",
                         "DCF PT", "Comps PT", "Blended PT", "Upside"]
    for ci, h in enumerate(scenario_headers, 2):
        style_header_sub(ws, r, ci, h)
    ws.row_dimensions[r].height = 18

    scenarios = [
        ("Bear",    A.WACC_BEAR, A.EXIT_BEAR, 6.0),
        ("Base ★",  A.WACC_BASE, A.EXIT_BASE, 8.0),
        ("Bull",    A.WACC_BULL, A.EXIT_BULL, 10.0),
    ]
    for ri, (sc, wacc_row, exit_row, qnx_mult) in enumerate(scenarios, r + 1):
        bold = "★" in sc
        style_label(ws, ri, 2, sc, bold=bold)
        wacc_f    = f"=Assumptions!$C${wacc_row}"
        exit_f    = f"=Assumptions!$C${exit_row}"
        qnx_str   = f"{qnx_mult:.1f}x"
        n         = len(PROJ_COLS)
        wacc_cell = f"Assumptions!$C${wacc_row}"
        dcf_f = (
            f"=({om}!$D${OM.UFCF}/(1+{wacc_cell})^1"
            f"+{om}!$E${OM.UFCF}/(1+{wacc_cell})^2"
            f"+{om}!$F${OM.UFCF}/(1+{wacc_cell})^3"
            f"+{om}!$G${OM.UFCF}/(1+{wacc_cell})^4"
            f"+{om}!$H${OM.UFCF}/(1+{wacc_cell})^5"
            f"+{om}!$H${OM.EBITDA}*Assumptions!$C${exit_row}/(1+{wacc_cell})^5"
            f"+{a}!$C${A.CASH})/{a}!$C${A.SHARES}"
        )
        comps_f = (
            f"=({om}!${comps_fy_col}${OM.QNX_REV}*{qnx_mult}"
            f"+{om}!${comps_fy_col}${OM.CSEC_REV}*Assumptions!$C${A.CSEC_COMPS}"
            f"+{a}!$C${A.CASH})/{a}!$C${A.SHARES}"
        )
        # Blended PT for this scenario (using scenario-specific weights)
        blend_f = (
            f"=Assumptions!$C${A.DCF_WEIGHT}*({om}!$D${OM.UFCF}/(1+{wacc_cell})^1"
            f"+{om}!$E${OM.UFCF}/(1+{wacc_cell})^2"
            f"+{om}!$F${OM.UFCF}/(1+{wacc_cell})^3"
            f"+{om}!$G${OM.UFCF}/(1+{wacc_cell})^4"
            f"+{om}!$H${OM.UFCF}/(1+{wacc_cell})^5"
            f"+{om}!$H${OM.EBITDA}*Assumptions!$C${exit_row}/(1+{wacc_cell})^5"
            f"+{a}!$C${A.CASH})/{a}!$C${A.SHARES}"
            f"+Assumptions!$C${A.COMPS_WEIGHT}*"
            f"({om}!${comps_fy_col}${OM.QNX_REV}*{qnx_mult}"
            f"+{om}!${comps_fy_col}${OM.CSEC_REV}*Assumptions!$C${A.CSEC_COMPS}"
            f"+{a}!$C${A.CASH})/{a}!$C${A.SHARES}"
        )
        upside_f = f"=({blend_f[1:]}-Assumptions!$C${A.CURR_PRICE})/Assumptions!$C${A.CURR_PRICE}"

        row_vals = [wacc_f, exit_f, qnx_str, dcf_f, comps_f, blend_f, upside_f]
        row_fmts = [FMT_PCT1, FMT_MULT, None, FMT_DOLLAR, FMT_DOLLAR, FMT_DOLLAR, FMT_PCT1]
        for ci, (v, fmt) in enumerate(zip(row_vals, row_fmts), 3):
            c = ws.cell(row=ri, column=ci, value=v)
            c.alignment = _align("center")
            if fmt:
                c.number_format = fmt
            if bold:
                c.font = _font(NAVY_FILL, bold=True)
        ws.row_dimensions[ri].height = 16

    ws.freeze_panes = "C5"


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    ws_cover  = wb.active
    ws_cover.title = "Cover"

    ws_assump = wb.create_sheet("Assumptions")
    ws_om     = wb.create_sheet("Operating Model")
    ws_dcf    = wb.create_sheet("DCF")
    ws_sens   = wb.create_sheet("Sensitivity")

    print("Building Cover...")
    build_cover(ws_cover)

    print("Building Assumptions...")
    build_assumptions(ws_assump)

    print("Building Operating Model...")
    build_operating_model(ws_om)

    print("Building DCF...")
    build_dcf(ws_dcf)

    print("Building Sensitivity...")
    build_sensitivity(ws_sens)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"✓  Model saved → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
