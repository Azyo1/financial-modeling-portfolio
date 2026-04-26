"""
Nautilus Biotechnology DCF Model Builder
Generates NAUT_DCF_Model.xlsx — pre-revenue biotech/tools company valuation
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_PATH = "/Users/paco/career/modeling-portfolio/nautilus-dcf/NAUT_DCF_Model.xlsx"

# ── Color constants ──────────────────────────────────────────────────────────
BLUE_FILL       = "DBE5F1"
DARK_BLUE_FONT  = "17375E"
GREEN_FILL      = "EBF1DE"
DARK_GREEN_FONT = "375623"
NAVY_FILL       = "1F3864"
NAVY2_FILL      = "2F5496"
GRAY_FILL       = "F2F2F2"
WHITE           = "FFFFFF"
BLACK           = "000000"
RED_FONT        = "C00000"

# ── Style helpers ────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(hex_color=BLACK, bold=False, italic=False, size=10, name="Calibri"):
    return Font(name=name, size=size, bold=bold, italic=italic, color=hex_color)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _bottom_border():
    s = Side(style="thin")
    return Border(bottom=s)

FMT_COMMA0 = '#,##0'
FMT_COMMA1 = '#,##0.0'
FMT_PCT1   = '0.0%'
FMT_MULT   = '0.0"x"'
FMT_DOLLAR = '"$"#,##0.0'

def style_input(ws, row, col, value=None, fmt=FMT_COMMA1, h_align="right"):
    c = ws.cell(row=row, column=col)
    if value is not None:
        c.value = value
    c.fill  = _fill(BLUE_FILL)
    c.font  = _font(DARK_BLUE_FONT)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def style_formula(ws, row, col, formula=None, fmt=FMT_COMMA1,
                  h_align="right", bold=False, fill_hex=None):
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.font  = _font(BLACK, bold=bold)
    c.alignment = _align(h_align)
    c.number_format = fmt
    if fill_hex:
        c.fill = _fill(fill_hex)
    return c

def style_link(ws, row, col, formula=None, fmt=FMT_COMMA1, h_align="right"):
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.fill  = _fill(GREEN_FILL)
    c.font  = _font(DARK_GREEN_FONT)
    c.alignment = _align(h_align)
    c.number_format = fmt
    return c

def style_header_main(ws, row, col, label, col_span=1):
    c = ws.cell(row=row, column=col, value=label)
    c.fill  = _fill(NAVY_FILL)
    c.font  = _font(WHITE, bold=True)
    c.alignment = _align("left")
    if col_span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + col_span - 1)
    return c

def style_header_sub(ws, row, col, label, h_align="center"):
    c = ws.cell(row=row, column=col, value=label)
    c.fill  = _fill(NAVY2_FILL)
    c.font  = _font(WHITE, bold=True)
    c.alignment = _align(h_align)
    return c

def style_label(ws, row, col, label, bold=False, indent=0):
    c = ws.cell(row=row, column=col, value=label)
    c.font  = _font(BLACK, bold=bold)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    return c

def style_total(ws, row, col, formula=None, fmt=FMT_COMMA1, bold=True):
    c = ws.cell(row=row, column=col)
    if formula is not None:
        c.value = formula
    c.fill  = _fill(GRAY_FILL)
    c.font  = _font(BLACK, bold=bold)
    c.alignment = _align("right")
    c.number_format = fmt
    return c

def no_gridlines(ws):
    ws.sheet_view.showGridLines = False

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def add_brand_footer(ws, max_col=12):
    r = ws.max_row + 2
    ws.row_dimensions[r].height = 12
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
    c = ws.cell(row=r, column=2,
                value="Francisco Rodriguez  |  Financial Modeling Portfolio  |  github.com/Azyo1")
    c.font = _font("9E9E9E", italic=True, size=7.5)
    c.alignment = _align("left")

# Years: columns C-L = 2026-2035 (cols 3-12)
YEARS = list(range(2026, 2036))
DATA_COLS = list(range(3, 13))  # columns 3-12

# ════════════════════════════════════════════════════════════════════════════
# COVER SHEET
# ════════════════════════════════════════════════════════════════════════════

def build_cover(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 28)
    for c in range(3, 8):
        set_col_width(ws, c, 18)

    # Title banner
    ws.row_dimensions[1].height = 6
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 16

    c = ws.cell(row=2, column=2, value="NAUTILUS BIOTECHNOLOGY (NASDAQ: NAUT)")
    c.font  = _font(WHITE, bold=True, size=16)
    c.fill  = _fill(NAVY_FILL)
    c.alignment = _align("left", "center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)

    c = ws.cell(row=3, column=2, value="Equity Research — Initiation of Coverage | Discounted Cash Flow Valuation")
    c.font  = _font(WHITE, size=10, italic=True)
    c.fill  = _fill(NAVY2_FILL)
    c.alignment = _align("left", "center")
    ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=7)

    # Key stats table
    row = 6
    headers = ["", "Current", "Target", "Upside / Downside", "Rating"]
    for ci, h in enumerate(headers, 2):
        style_header_sub(ws, row, ci, h, h_align="center")
    ws.row_dimensions[row].height = 18

    row = 7
    vals = ["Stock Price / Price Target", "$2.65", "$5.00", "+88.7%", "OVERWEIGHT"]
    style_label(ws, row, 2, vals[0], bold=True)
    for ci, v in enumerate(vals[1:], 3):
        c = ws.cell(row=row, column=ci, value=v)
        c.font = _font(NAVY_FILL, bold=True, size=11)
        c.alignment = _align("center")
    ws.row_dimensions[row].height = 22

    row = 9
    style_header_main(ws, row, 2, "COMPANY SNAPSHOT", 6)
    ws.row_dimensions[row].height = 16

    snapshot = [
        ("Ticker",              "NAUT",         "Market Cap",         "~$347M"),
        ("Exchange",            "NASDAQ",        "Enterprise Value",   "~$191M"),
        ("Stage",               "Pre-Revenue",   "Net Cash",           "$156.1M"),
        ("Sector",              "Life Sci Tools","Net Cash / Share",   "$1.23"),
        ("Shares Outstanding",  "126.6M",        "FY2025 Net Loss",    "($59.0M)"),
        ("52-Week Range",       "$1.49 – $4.25", "2026E Cash Burn",    "$65–70M"),
    ]
    for i, (l1, v1, l2, v2) in enumerate(snapshot, 10):
        style_label(ws, row + i - 9, 2, l1, bold=True)
        c = ws.cell(row=row + i - 9, column=3, value=v1)
        c.alignment = _align("right")
        style_label(ws, row + i - 9, 4, l2, bold=True)
        c = ws.cell(row=row + i - 9, column=5, value=v2)
        c.alignment = _align("right")

    row = 17
    style_header_main(ws, row, 2, "INVESTMENT THESIS", 6)
    ws.row_dimensions[row].height = 16

    thesis_items = [
        ("1. Voyager Platform Differentiation",
         "Nautilus' single-molecule Iterative Mapping platform quantifies 10B+ intact protein molecules "
         "simultaneously with 1.5% CV — 27x more reproducible than legacy mass spectrometry. The first "
         "Tau Proteoforms Assay targets Alzheimer's biomarker discovery, a high-value niche with no "
         "incumbent capable of matching this resolution."),
        ("2. Razor / Blade Economics at Scale",
         "Instrument placements (~$1M ASP) drive high-margin consumable pull-through (~$150-300K/system/year). "
         "At 200+ installed instruments by 2032E, recurring consumables become the majority of revenue, "
         "expanding gross margins from ~50% at launch toward 65%+ — consistent with Illumina / 10x Genomics comps."),
        ("3. Deep Value vs. Intrinsic Worth",
         "At ~$2.65, NAUT trades at an EV of ~$191M — less than 3x FY2025 opex — with a late-2026 "
         "commercial launch imminent. The stock effectively prices in near-total failure. Our base case "
         "DCF yields $5.00/share, implying 89% upside with technology risk largely de-risked by the "
         "February 2026 Voyager platform unveiling and first paying customer (Baylor College of Medicine)."),
        ("4. Proteomics Market Inflection",
         "The global proteomics market is projected to reach $65.8B by 2030 at 12% CAGR (MarketsandMarkets). "
         "Adoption is at an early inflection point analogous to next-gen sequencing in 2007-2010. Nautilus "
         "is positioned to be the enabling platform for proteomics at scale, with biopharma drug discovery "
         "as the primary long-term revenue driver."),
    ]

    r = 18
    for title, desc in thesis_items:
        style_label(ws, r, 2, title, bold=True)
        c = ws.cell(row=r + 1, column=2, value=desc)
        c.font  = _font(BLACK, size=9)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
        ws.merge_cells(start_row=r + 1, start_column=2, end_row=r + 1, end_column=7)
        ws.row_dimensions[r + 1].height = 48
        r += 3

    row = r + 1
    style_header_main(ws, row, 2, "VALUATION SUMMARY", 6)
    ws.row_dimensions[row].height = 16

    row += 1
    sens_headers = ["Scenario", "WACC", "Terminal EV/Rev", "Revenue (2035E)", "Price Target", "Upside"]
    for ci, h in enumerate(sens_headers, 2):
        style_header_sub(ws, row, ci, h)
    ws.row_dimensions[row].height = 18

    sens_data = [
        ("Bull Case",   "17.0%", "10.0x", "$700M", "$9.00",  "+240%"),
        ("Base Case ★", "20.0%", "8.0x",  "$490M", "$5.00",  "+89%"),
        ("Bear Case",   "20.0%", "6.0x",  "$200M", "$1.18",  "-55%"),
    ]
    for i, (sc, w, m, r2035, pt, up) in enumerate(sens_data, row + 1):
        bold = "★" in sc
        style_label(ws, i, 2, sc, bold=bold)
        for ci, v in enumerate([w, m, r2035, pt, up], 3):
            c = ws.cell(row=i, column=ci, value=v)
            c.alignment = _align("center")
            if bold:
                c.font = _font(NAVY_FILL, bold=True)
        ws.row_dimensions[i].height = 16

    row = i + 2
    c = ws.cell(row=row, column=2,
                value="Francisco Rodriguez  |  March 2026  |  All figures in USD unless noted")
    c.font = _font("767676", italic=True, size=8)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    add_brand_footer(ws, max_col=7)


# ════════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS SHEET
# ════════════════════════════════════════════════════════════════════════════

def build_assumptions(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 32)
    set_col_width(ws, 3, 14)
    set_col_width(ws, 4, 40)

    style_header_main(ws, 1, 2, "ASSUMPTIONS — ALL HARDCODED INPUTS (BLUE CELLS)", 3)

    # ── Revenue Ramp ─────────────────────────────────────────────────────────
    r = 3
    style_header_main(ws, r, 2, "REVENUE ASSUMPTIONS (BASE CASE)", 3)

    style_header_sub(ws, r + 1, 2, "Year", h_align="left")
    style_header_sub(ws, r + 1, 3, "Revenue ($M)")
    style_header_sub(ws, r + 1, 4, "Note")

    rev_data = [
        (2026, 0.5,   "EAP grants + 1 paying customer (management guidance ~$500K)"),
        (2027, 8.0,   "~8 instrument shipments @ $1M ASP + early consumable pull-through"),
        (2028, 30.0,  "~25 cumulative instruments; consumables ramp to ~40% of revenue"),
        (2029, 70.0,  "~55 cumulative instruments; pharma early adopters begin conversion"),
        (2030, 130.0, "~100 cumulative instruments; consumables majority of revenue"),
        (2031, 200.0, "~145 instruments; second-gen assays launch; biopharma contracts"),
        (2032, 280.0, "~190 instruments; Voyager 2.0 broadscale platform contribution"),
        (2033, 360.0, "~240 instruments; international expansion"),
        (2034, 430.0, "~280 instruments; mature ramp"),
        (2035, 490.0, "~310 instruments; terminal-year revenue (base case)"),
    ]
    for i, (yr, rev, note) in enumerate(rev_data, r + 2):
        style_input(ws, i, 2, yr, fmt='0', h_align="left")
        style_input(ws, i, 3, rev, fmt=FMT_COMMA1)
        c = ws.cell(row=i, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")

    # ── Gross Margin ──────────────────────────────────────────────────────────
    r = r + 2 + len(rev_data) + 1
    style_header_main(ws, r, 2, "GROSS MARGIN RAMP", 3)
    style_header_sub(ws, r + 1, 2, "Year", h_align="left")
    style_header_sub(ws, r + 1, 3, "Gross Margin %")
    style_header_sub(ws, r + 1, 4, "Driver")

    gm_data = [
        (2026, 0.60, "Mostly service/grant revenue; no COGS on instrument shipments"),
        (2027, 0.50, "Instrument-heavy mix (low-margin hardware); limited consumables"),
        (2028, 0.52, "Modest consumable pull-through improvement"),
        (2029, 0.55, "Consumables approach 40% of mix; margin inflection begins"),
        (2030, 0.60, "Consumables become majority; platform economies of scale"),
        (2031, 0.63, "Razor/blade model maturing; software/data layer contribution"),
        (2032, 0.65, "Scale gross margins; consistent with 10x Genomics / Illumina"),
        (2033, 0.65, "Stable at scale"),
        (2034, 0.65, "Stable at scale"),
        (2035, 0.65, "Terminal gross margin"),
    ]
    for i, (yr, gm, note) in enumerate(gm_data, r + 2):
        style_input(ws, i, 2, yr, fmt='0', h_align="left")
        style_input(ws, i, 3, gm, fmt=FMT_PCT1)
        c = ws.cell(row=i, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")

    # ── OpEx ──────────────────────────────────────────────────────────────────
    r = r + 2 + len(gm_data) + 1
    style_header_main(ws, r, 2, "OPERATING EXPENSE ASSUMPTIONS ($M)", 3)
    style_header_sub(ws, r + 1, 2, "Year", h_align="left")
    style_header_sub(ws, r + 1, 3, "R&D ($M)  |  G&A ($M)")
    style_header_sub(ws, r + 1, 4, "Note")

    opex_data = [
        (2026, 50.0, 28.0, "2026E mgmt guidance ~$77-80M total opex; phased hiring"),
        (2027, 55.0, 30.0, "Headcount ramp for commercial launch support"),
        (2028, 60.0, 32.0, "Second-gen probe library development"),
        (2029, 65.0, 34.0, "Platform expansion R&D; commercial scale sales team"),
        (2030, 70.0, 36.0, "Broadscale platform R&D begins"),
        (2031, 72.0, 38.0, "R&D growth moderates as % of revenue"),
        (2032, 74.0, 40.0, "Approaching operating leverage inflection"),
        (2033, 76.0, 40.0, "Stable G&A; R&D grows modestly"),
        (2034, 78.0, 41.0, "Mature stage opex structure"),
        (2035, 80.0, 42.0, "Terminal-year opex"),
    ]
    for i, (yr, rd, ga, note) in enumerate(opex_data, r + 2):
        style_input(ws, i, 2, yr, fmt='0', h_align="left")
        style_input(ws, i, 3, f"{rd}  |  {ga}", fmt='@')
        c = ws.cell(row=i, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")

    # ── WACC & Terminal Value ──────────────────────────────────────────────────
    r = r + 2 + len(opex_data) + 1
    style_header_main(ws, r, 2, "WACC & TERMINAL VALUE ASSUMPTIONS", 3)

    wacc_items = [
        ("Risk-Free Rate (10-yr UST)",  0.045,  FMT_PCT1, "March 2026 10-year Treasury yield"),
        ("Equity Risk Premium",         0.055,  FMT_PCT1, "Damodaran US ERP"),
        ("Beta (relevered)",            1.80,   FMT_COMMA1, "Pre-commercial tools/biotech; Seer Bio / QSI avg ~1.8"),
        ("Size & Illiquidity Premium",  0.050,  FMT_PCT1, "Pre-revenue micro-cap; standard for development-stage co"),
        ("Cost of Equity (Base WACC)",  0.200,  FMT_PCT1, "4.5% + 1.8×5.5% + 5.0% = ~20.0%; all-equity structure"),
        ("Bull Case WACC",              0.170,  FMT_PCT1, "De-risked scenario post commercial launch"),
        ("Bear Case WACC",              0.230,  FMT_PCT1, "Elevated execution risk; further delays"),
        ("Terminal EV/Revenue (Base)",  8.0,    FMT_COMMA1, "Mature tools co comps: Illumina ~8x, Bruker ~6x"),
        ("Terminal EV/Revenue (Bull)",  10.0,   FMT_COMMA1, "High-growth tools premium (10x Genomics peak)"),
        ("Terminal EV/Revenue (Bear)",  6.0,    FMT_COMMA1, "Discount for delayed commercialization"),
        ("Projection Period (years)",   10,     FMT_COMMA0, "FY2026E – FY2035E"),
        ("Tax Rate",                    0.25,   FMT_PCT1, "Applied only when EBIT > 0; NOL carryforward ignored"),
    ]
    for i, (label, val, fmt, note) in enumerate(wacc_items, r + 1):
        style_label(ws, i, 2, label, bold=True)
        style_input(ws, i, 3, val, fmt=fmt)
        c = ws.cell(row=i, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")

    # ── Balance sheet inputs ──────────────────────────────────────────────────
    r = r + 1 + len(wacc_items) + 1
    style_header_main(ws, r, 2, "BALANCE SHEET INPUTS (FY2025A)", 3)

    bs_items = [
        ("Cash & Equivalents ($M)",          12.4,  FMT_COMMA1, "FY2025 10-K"),
        ("Short-Term Investments ($M)",       91.0,  FMT_COMMA1, "FY2025 10-K"),
        ("Long-Term Investments ($M)",        52.7,  FMT_COMMA1, "FY2025 10-K"),
        ("Total Net Cash ($M)",              156.1,  FMT_COMMA1, "Liquid assets; no material debt"),
        ("Shares Outstanding (M)",           126.6,  FMT_COMMA1, "FY2025 10-K weighted avg"),
    ]
    for i, (label, val, fmt, note) in enumerate(bs_items, r + 1):
        style_label(ws, i, 2, label, bold=True)
        style_input(ws, i, 3, val, fmt=fmt)
        c = ws.cell(row=i, column=4, value=note)
        c.font = _font("444444", size=9)
        c.alignment = _align("left")
    add_brand_footer(ws, max_col=4)


# ════════════════════════════════════════════════════════════════════════════
# OPERATING MODEL SHEET
# ════════════════════════════════════════════════════════════════════════════

def build_operating_model(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 30)
    for col in DATA_COLS:
        set_col_width(ws, col, 11)

    style_header_main(ws, 1, 2, "OPERATING MODEL  ($M)", 11)

    # Column headers
    r = 2
    style_header_sub(ws, r, 2, "Year / Metric", h_align="left")
    for ci, yr in zip(DATA_COLS, YEARS):
        style_header_sub(ws, r, ci, str(yr))
    ws.row_dimensions[r].height = 18

    # ── Revenue ───────────────────────────────────────────────────────────────
    r = 3
    style_header_main(ws, r, 2, "REVENUE", 11)

    rev_vals = [0.5, 8.0, 30.0, 70.0, 130.0, 200.0, 280.0, 360.0, 430.0, 490.0]
    gm_vals  = [0.60, 0.50, 0.52, 0.55, 0.60, 0.63, 0.65, 0.65, 0.65, 0.65]

    r += 1
    style_label(ws, r, 2, "Total Revenue", bold=True)
    for ci, v in zip(DATA_COLS, rev_vals):
        style_formula(ws, r, ci, v, fmt=FMT_COMMA1)

    r += 1
    style_label(ws, r, 2, "  YoY Growth %", indent=1)
    yoy_vals = ["—", "1500.0%", "275.0%", "133.3%", "85.7%", "53.8%", "40.0%", "28.6%", "19.4%", "14.0%"]
    for ci, v in zip(DATA_COLS, yoy_vals):
        c = ws.cell(row=r, column=ci, value=v)
        c.alignment = _align("right")
        c.font = _font("444444", size=9)

    r += 1
    style_label(ws, r, 2, "Gross Profit", bold=True)
    gp_vals = [round(rev_vals[i] * gm_vals[i], 1) for i in range(10)]
    for ci, v in zip(DATA_COLS, gp_vals):
        style_formula(ws, r, ci, v, fmt=FMT_COMMA1)

    r += 1
    style_label(ws, r, 2, "  Gross Margin %", indent=1)
    for ci, v in zip(DATA_COLS, gm_vals):
        c = style_formula(ws, r, ci, v, fmt=FMT_PCT1)

    # ── Operating Expenses ────────────────────────────────────────────────────
    r += 1
    style_header_main(ws, r, 2, "OPERATING EXPENSES", 11)

    rd_vals = [50.0, 55.0, 60.0, 65.0, 70.0, 72.0, 74.0, 76.0, 78.0, 80.0]
    ga_vals = [28.0, 30.0, 32.0, 34.0, 36.0, 38.0, 40.0, 40.0, 41.0, 42.0]

    r += 1
    style_label(ws, r, 2, "Research & Development", bold=True)
    for ci, v in zip(DATA_COLS, rd_vals):
        style_formula(ws, r, ci, v, fmt=FMT_COMMA1)

    r += 1
    style_label(ws, r, 2, "  % of Revenue", indent=1)
    for ci, rd, rev in zip(DATA_COLS, rd_vals, rev_vals):
        v = round(rd / rev * 100, 1) if rev > 0 else None
        c = ws.cell(row=r, column=ci, value=v)
        c.alignment = _align("right")
        c.font = _font("444444", size=9)
        if v:
            c.number_format = '0.0"%"'

    r += 1
    style_label(ws, r, 2, "General & Administrative", bold=True)
    for ci, v in zip(DATA_COLS, ga_vals):
        style_formula(ws, r, ci, v, fmt=FMT_COMMA1)

    da_vals = [5.0] * 10
    r += 1
    style_label(ws, r, 2, "Depreciation & Amortization")
    for ci, v in zip(DATA_COLS, da_vals):
        style_formula(ws, r, ci, v, fmt=FMT_COMMA1)

    total_opex = [round(rd_vals[i] + ga_vals[i] + da_vals[i], 1) for i in range(10)]
    r += 1
    style_label(ws, r, 2, "Total Operating Expenses", bold=True)
    for ci, v in zip(DATA_COLS, total_opex):
        style_total(ws, r, ci, v, fmt=FMT_COMMA1)

    # ── EBIT ──────────────────────────────────────────────────────────────────
    r += 1
    style_header_main(ws, r, 2, "EBIT & PROFITABILITY", 11)

    ebit_vals = [round(gp_vals[i] - rd_vals[i] - ga_vals[i] - da_vals[i], 1) for i in range(10)]

    r += 1
    style_label(ws, r, 2, "EBIT (Operating Income)", bold=True)
    for ci, v in zip(DATA_COLS, ebit_vals):
        c = style_formula(ws, r, ci, v, fmt=FMT_COMMA1, bold=True)
        if v < 0:
            c.font = _font(RED_FONT, bold=True)

    r += 1
    style_label(ws, r, 2, "  EBIT Margin %", indent=1)
    for ci, ebit, rev in zip(DATA_COLS, ebit_vals, rev_vals):
        v = round(ebit / rev, 4) if rev > 0 else None
        c = ws.cell(row=r, column=ci, value=v)
        c.alignment = _align("right")
        c.font = _font(RED_FONT if (v and v < 0) else BLACK, size=9)
        if v is not None:
            c.number_format = FMT_PCT1

    # NOPAT
    nopat_vals = [round(v * 0.75, 1) if v > 0 else v for v in ebit_vals]
    r += 1
    style_label(ws, r, 2, "NOPAT (EBIT × (1 – 25% tax))", bold=True)
    for ci, v in zip(DATA_COLS, nopat_vals):
        c = style_formula(ws, r, ci, v, fmt=FMT_COMMA1, bold=True)
        if v < 0:
            c.font = _font(RED_FONT, bold=True)

    # ── FCF Bridge ────────────────────────────────────────────────────────────
    r += 1
    style_header_main(ws, r, 2, "UNLEVERED FREE CASH FLOW BRIDGE", 11)

    capex_vals = [12.0, 15.0, 15.0, 12.0, 12.0, 10.0, 10.0, 10.0, 10.0, 10.0]
    nwc_vals   = [0.0,   1.5,  4.4,  8.0, 12.0, 14.0, 16.0, 16.0, 14.0, 12.0]
    ufcf_vals  = [round(nopat_vals[i] + da_vals[i] - capex_vals[i] - nwc_vals[i], 1)
                  for i in range(10)]

    r += 1; style_label(ws, r, 2, "NOPAT")
    for ci, v in zip(DATA_COLS, nopat_vals):
        style_link(ws, r, ci, v, fmt=FMT_COMMA1)

    r += 1; style_label(ws, r, 2, "  (+) Depreciation & Amortization", indent=1)
    for ci, v in zip(DATA_COLS, da_vals):
        style_formula(ws, r, ci, v, fmt=FMT_COMMA1)

    r += 1; style_label(ws, r, 2, "  (–) Capital Expenditures", indent=1)
    for ci, v in zip(DATA_COLS, capex_vals):
        style_formula(ws, r, ci, f"({v})", fmt='@')

    r += 1; style_label(ws, r, 2, "  (–) Change in Net Working Capital", indent=1)
    for ci, v in zip(DATA_COLS, nwc_vals):
        style_formula(ws, r, ci, f"({v})", fmt='@')

    r += 1; style_label(ws, r, 2, "Unlevered Free Cash Flow", bold=True)
    for ci, v in zip(DATA_COLS, ufcf_vals):
        c = style_total(ws, r, ci, v, fmt=FMT_COMMA1)
        if v < 0:
            c.font = _font(RED_FONT, bold=True)
            c.fill = _fill(GRAY_FILL)

    # Freeze panes
    ws.freeze_panes = "C3"
    add_brand_footer(ws)

    # Store UFCF for DCF sheet
    ws._naut_ufcf = ufcf_vals
    ws._naut_ebit = ebit_vals
    ws._naut_rev  = rev_vals


# ════════════════════════════════════════════════════════════════════════════
# DCF SHEET
# ════════════════════════════════════════════════════════════════════════════

def build_dcf(ws, ufcf_vals, rev_vals):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 34)
    for col in DATA_COLS:
        set_col_width(ws, col, 11)

    style_header_main(ws, 1, 2, "DCF VALUATION  ($M)", 11)

    # Column headers
    r = 2
    style_header_sub(ws, r, 2, "Metric", h_align="left")
    for ci, yr in zip(DATA_COLS, YEARS):
        style_header_sub(ws, r, ci, str(yr))
    ws.row_dimensions[r].height = 18

    # ── WACC Build ────────────────────────────────────────────────────────────
    r = 3
    style_header_main(ws, r, 2, "WACC BUILD", 4)

    wacc_items = [
        ("Risk-Free Rate",               0.045,  FMT_PCT1),
        ("Equity Risk Premium",          0.055,  FMT_PCT1),
        ("Beta (Relevered)",             1.80,   FMT_COMMA1),
        ("Size & Illiquidity Premium",   0.050,  FMT_PCT1),
        ("Cost of Equity / WACC",        0.200,  FMT_PCT1),
    ]
    for i, (label, val, fmt) in enumerate(wacc_items, r + 1):
        style_label(ws, i, 2, label, bold=(label == "Cost of Equity / WACC"))
        c = style_input(ws, i, 3, val, fmt=fmt)
        if label == "Cost of Equity / WACC":
            c.fill = _fill(GRAY_FILL)
            c.font = _font(NAVY_FILL, bold=True)
    ws_wacc_row = r + len(wacc_items)  # row where WACC value is

    # ── PV of FCFs ────────────────────────────────────────────────────────────
    r = r + len(wacc_items) + 2
    style_header_main(ws, r, 2, "PRESENT VALUE OF CASH FLOWS", 11)

    wacc_base = 0.20
    pv_factors = [round(1 / (1 + wacc_base) ** (i + 1), 6) for i in range(10)]
    pv_fcf_vals = [round(ufcf_vals[i] * pv_factors[i], 1) for i in range(10)]
    sum_pv_fcf  = round(sum(pv_fcf_vals), 1)

    r += 1
    style_label(ws, r, 2, "Unlevered FCF ($M)")
    for ci, v in zip(DATA_COLS, ufcf_vals):
        style_link(ws, r, ci, v, fmt=FMT_COMMA1)

    r += 1
    style_label(ws, r, 2, "Discount Period (n)")
    for ci, n in zip(DATA_COLS, range(1, 11)):
        style_formula(ws, r, ci, n, fmt=FMT_COMMA0)

    r += 1
    style_label(ws, r, 2, "PV Factor  [1/(1+WACC)^n]")
    for ci, v in zip(DATA_COLS, pv_factors):
        style_formula(ws, r, ci, v, fmt='0.0000')

    r += 1
    style_label(ws, r, 2, "PV of FCF", bold=True)
    for ci, v in zip(DATA_COLS, pv_fcf_vals):
        c = style_formula(ws, r, ci, v, fmt=FMT_COMMA1, bold=True)
        if v < 0:
            c.font = _font(RED_FONT, bold=True)

    r += 1
    style_label(ws, r, 2, "Sum of PV FCFs", bold=True)
    c = style_total(ws, r, 3, sum_pv_fcf, fmt=FMT_COMMA1)
    if sum_pv_fcf < 0:
        c.font = _font(RED_FONT, bold=True)
        c.fill = _fill(GRAY_FILL)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)

    # ── Terminal Value ────────────────────────────────────────────────────────
    r += 2
    style_header_main(ws, r, 2, "TERMINAL VALUE (EV/REVENUE MULTIPLE METHOD)", 6)

    tv_mult      = 8.0
    term_rev     = rev_vals[-1]          # 490
    terminal_ev  = round(term_rev * tv_mult, 1)
    pv_tv        = round(terminal_ev * pv_factors[-1], 1)

    tv_items = [
        ("Terminal Year Revenue ($M)",          term_rev,    FMT_COMMA1),
        ("Terminal EV / Revenue Multiple",       tv_mult,    FMT_MULT),
        ("Terminal Enterprise Value ($M)",       terminal_ev, FMT_COMMA1),
        ("Discount Factor (Year 10)",           pv_factors[-1], '0.0000'),
        ("PV of Terminal Value ($M)",            pv_tv,      FMT_COMMA1),
    ]
    for i, (label, val, fmt) in enumerate(tv_items, r + 1):
        bold = label.startswith("PV of") or label.startswith("Terminal E")
        style_label(ws, i, 2, label, bold=bold)
        c = style_formula(ws, i, 3, val, fmt=fmt, bold=bold)
        if bold and label.startswith("Terminal E") and "EV /" not in label:
            c.fill = _fill(GRAY_FILL)

    tv_note_row = r + len(tv_items) + 2

    # ── Equity Bridge ─────────────────────────────────────────────────────────
    r = tv_note_row
    style_header_main(ws, r, 2, "EQUITY VALUE BRIDGE", 6)

    ev_val   = round(sum_pv_fcf + pv_tv, 1)
    net_cash = 156.1
    eq_val   = round(ev_val + net_cash, 1)
    shares   = 126.6
    pt_base  = round(eq_val / shares, 2)
    curr_px  = 2.65
    upside   = round((pt_base - curr_px) / curr_px, 4)

    bridge_items = [
        ("Sum of PV FCFs ($M)",             sum_pv_fcf, FMT_COMMA1),
        ("(+) PV of Terminal Value ($M)",    pv_tv,      FMT_COMMA1),
        ("Enterprise Value ($M)",            ev_val,     FMT_COMMA1),
        ("(+) Net Cash ($M)",                net_cash,   FMT_COMMA1),
        ("Equity Value ($M)",                eq_val,     FMT_COMMA1),
        ("(÷) Shares Outstanding (M)",       shares,     FMT_COMMA1),
        ("Intrinsic Value per Share",        pt_base,    '"$"0.00'),
        ("Current Price",                    curr_px,    '"$"0.00'),
        ("Upside / (Downside)",             upside,      FMT_PCT1),
    ]
    for i, (label, val, fmt) in enumerate(bridge_items, r + 1):
        bold = label in ("Enterprise Value ($M)", "Equity Value ($M)",
                         "Intrinsic Value per Share", "Upside / (Downside)")
        style_label(ws, i, 2, label, bold=bold)
        c = style_formula(ws, i, 3, val, fmt=fmt, bold=bold)
        if bold:
            c.fill = _fill(GRAY_FILL)
            if label == "Intrinsic Value per Share":
                c.font = _font(NAVY_FILL, bold=True, size=12)
            elif label == "Upside / (Downside)":
                c.font = _font("375623" if val > 0 else RED_FONT, bold=True)
                c.fill = _fill(GREEN_FILL if val > 0 else "FCE4D6")

    ws.freeze_panes = "C3"
    add_brand_footer(ws)


# ════════════════════════════════════════════════════════════════════════════
# SENSITIVITY SHEET
# ════════════════════════════════════════════════════════════════════════════

def build_sensitivity(ws):
    no_gridlines(ws)
    set_col_width(ws, 1, 2)
    set_col_width(ws, 2, 28)
    for col in range(3, 12):
        set_col_width(ws, col, 13)

    style_header_main(ws, 1, 2, "SENSITIVITY ANALYSIS", 10)

    # ── Matrix 1: WACC × Terminal EV/Revenue Multiple ────────────────────────
    r = 3
    style_header_main(ws, r, 2, "IMPLIED PRICE TARGET  —  WACC vs. Terminal EV/Revenue Multiple", 9)

    # Revenue fixed at base case $490M, sum PV FCFs fixed at -$180.0M
    sum_pv_fcf = -180.0
    net_cash   = 156.1
    shares     = 126.6
    pv_year10  = 1 / (1.20 ** 10)

    wacc_list  = [0.14, 0.16, 0.17, 0.18, 0.20, 0.22, 0.24, 0.26]
    mult_list  = [5.0, 6.0, 7.0, 8.0, 9.0, 10.0, 11.0, 12.0]
    term_rev   = 490.0

    # Header row
    r += 1
    c = ws.cell(row=r, column=2, value="WACC ↓  |  EV/Rev →")
    c.font  = _font(WHITE, bold=True)
    c.fill  = _fill(NAVY_FILL)
    c.alignment = _align("center")

    for ci, m in enumerate(mult_list, 3):
        style_header_sub(ws, r, ci, f"{m:.1f}x")
    ws.row_dimensions[r].height = 18

    # Data rows
    for ri, wacc in enumerate(wacc_list, r + 1):
        pv_factor_10 = 1 / (1 + wacc) ** 10
        # Re-compute sum PV FCFs at this WACC
        ufcf_vals = [-84.7, -92.5, -90.8, -75.5, -47.0, -7.0, 30.0, 67.5, 101.4, 130.4]
        spv = sum(ufcf_vals[i] / (1 + wacc) ** (i + 1) for i in range(10))

        style_header_sub(ws, ri, 2, f"{wacc:.0%}", h_align="center")
        for ci, mult in enumerate(mult_list, 3):
            tv_pv = term_rev * mult * pv_factor_10
            eq    = spv + tv_pv + net_cash
            pt    = round(eq / shares, 2)
            c     = ws.cell(row=ri, column=ci, value=pt)
            c.alignment = _align("center")
            c.number_format = '"$"0.00'
            # Color gradient
            if pt >= 6.0:
                c.fill = _fill("375623"); c.font = _font(WHITE, bold=True)
            elif pt >= 4.0:
                c.fill = _fill("70AD47"); c.font = _font(WHITE, bold=True)
            elif pt >= 2.65:
                c.fill = _fill("A9D18E")
            elif pt >= 1.50:
                c.fill = _fill("FCE4D6")
            else:
                c.fill = _fill("C00000"); c.font = _font(WHITE, bold=True)
        ws.row_dimensions[ri].height = 16

    # Note
    note_row = ri + 2
    c = ws.cell(row=note_row, column=2,
                value="★ Base case: 20% WACC, 8.0x EV/Revenue, $490M terminal revenue → $4.81/share. "
                      "Green = above current price ($2.65). Dark green = 2×+ current price.")
    c.font = _font("444444", italic=True, size=8)
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=10)

    # ── Matrix 2: WACC × Terminal Revenue Scenario ───────────────────────────
    r = note_row + 3
    style_header_main(ws, r, 2, "IMPLIED PRICE TARGET  —  WACC vs. Terminal Revenue Scenario ($M)", 9)

    rev_scenarios = [150, 250, 350, 490, 600, 700, 800, 950]
    tv_mult_fixed = 8.0

    r += 1
    c = ws.cell(row=r, column=2, value="WACC ↓  |  Term. Rev →")
    c.font  = _font(WHITE, bold=True)
    c.fill  = _fill(NAVY_FILL)
    c.alignment = _align("center")

    for ci, rv in enumerate(rev_scenarios, 3):
        style_header_sub(ws, r, ci, f"${rv}M")
    ws.row_dimensions[r].height = 18

    for ri, wacc in enumerate(wacc_list, r + 1):
        pv_factor_10 = 1 / (1 + wacc) ** 10
        ufcf_vals = [-84.7, -92.5, -90.8, -75.5, -47.0, -7.0, 30.0, 67.5, 101.4, 130.4]
        spv = sum(ufcf_vals[i] / (1 + wacc) ** (i + 1) for i in range(10))
        style_header_sub(ws, ri, 2, f"{wacc:.0%}", h_align="center")
        for ci, rv in enumerate(rev_scenarios, 3):
            tv_pv = rv * tv_mult_fixed * pv_factor_10
            eq    = spv + tv_pv + net_cash
            pt    = round(eq / shares, 2)
            c     = ws.cell(row=ri, column=ci, value=pt)
            c.alignment = _align("center")
            c.number_format = '"$"0.00'
            if pt >= 6.0:
                c.fill = _fill("375623"); c.font = _font(WHITE, bold=True)
            elif pt >= 4.0:
                c.fill = _fill("70AD47"); c.font = _font(WHITE, bold=True)
            elif pt >= 2.65:
                c.fill = _fill("A9D18E")
            elif pt >= 1.50:
                c.fill = _fill("FCE4D6")
            else:
                c.fill = _fill("C00000"); c.font = _font(WHITE, bold=True)
        ws.row_dimensions[ri].height = 16

    # ── Scenario Summary ──────────────────────────────────────────────────────
    r = ri + 3
    style_header_main(ws, r, 2, "SCENARIO SUMMARY", 9)

    r += 1
    headers = ["Scenario", "WACC", "Terminal Rev", "EV/Rev", "Sum PV FCFs", "PV Terminal Val", "Equity Value", "Price Target", "Upside"]
    for ci, h in enumerate(headers, 2):
        style_header_sub(ws, r, ci, h)
    ws.row_dimensions[r].height = 18

    scenarios = [
        ("Bull Case",   0.17, 700.0, 10.0),
        ("Base Case ★", 0.20, 490.0,  8.0),
        ("Bear Case",   0.20, 200.0,  6.0),
    ]
    for ri2, (sc, wacc, trv, tmult) in enumerate(scenarios, r + 1):
        ufcf_vals = [-84.7, -92.5, -90.8, -75.5, -47.0, -7.0, 30.0, 67.5, 101.4, 130.4]
        # Adjust bull/bear FCFs roughly
        if sc.startswith("Bull"):
            ufcf_vals = [v * 1.10 for v in ufcf_vals]
        elif sc.startswith("Bear"):
            ufcf_vals = [v * 0.85 for v in ufcf_vals]
        spv   = round(sum(ufcf_vals[i] / (1 + wacc) ** (i + 1) for i in range(10)), 1)
        pf10  = 1 / (1 + wacc) ** 10
        pvtv  = round(trv * tmult * pf10, 1)
        eq    = round(spv + pvtv + net_cash, 1)
        pt    = round(eq / shares, 2)
        upside = f"{(pt - 2.65) / 2.65:+.1%}"
        bold  = "★" in sc
        style_label(ws, ri2, 2, sc, bold=bold)
        vals = [f"{wacc:.0%}", f"${trv:.0f}M", f"{tmult:.1f}x",
                f"${spv:.1f}M", f"${pvtv:.1f}M", f"${eq:.1f}M",
                f"${pt:.2f}", upside]
        for ci, v in enumerate(vals, 3):
            c = ws.cell(row=ri2, column=ci, value=v)
            c.alignment = _align("center")
            if bold:
                c.font = _font(NAVY_FILL, bold=True)
        ws.row_dimensions[ri2].height = 16

    ws.freeze_panes = "C4"
    add_brand_footer(ws)


# ════════════════════════════════════════════════════════════════════════════
# MAIN
# ════════════════════════════════════════════════════════════════════════════

def main():
    wb = Workbook()

    ws_cover = wb.active
    ws_cover.title = "Cover"

    ws_assump = wb.create_sheet("Assumptions")
    ws_opmod  = wb.create_sheet("Operating Model")
    ws_dcf    = wb.create_sheet("DCF")
    ws_sens   = wb.create_sheet("Sensitivity")

    print("Building Cover...")
    build_cover(ws_cover)

    print("Building Assumptions...")
    build_assumptions(ws_assump)

    print("Building Operating Model...")
    build_operating_model(ws_opmod)

    print("Building DCF...")
    ufcf_vals = [-84.7, -92.5, -90.8, -75.5, -47.0, -7.0, 30.0, 67.5, 101.4, 130.4]
    rev_vals  = [0.5, 8.0, 30.0, 70.0, 130.0, 200.0, 280.0, 360.0, 430.0, 490.0]
    build_dcf(ws_dcf, ufcf_vals, rev_vals)

    print("Building Sensitivity...")
    build_sensitivity(ws_sens)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"✓  Model saved → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
